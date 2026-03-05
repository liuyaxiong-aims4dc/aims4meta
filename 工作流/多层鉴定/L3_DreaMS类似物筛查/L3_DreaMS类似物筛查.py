#!/usr/bin/env python3
"""
L3: DreaMS 类似物筛查（纯 Embedding 语义匹配）

策略：
- 不设母离子偏差、子离子匹配数、质量偏差约束
- 纯 embedding cosine 相似度筛查
- 候选库：L1 真实库 + L2 模拟库合并的全部 embedding
- 碎片匹配：仍然计算并记录（仅作为参考信息，不作为过滤条件）

输入：L2 未鉴定的样品 embedding
输出：L3_results.csv + L3_results.xlsx + L3_analog.msp + L3_remaining.msp
"""

import os
import sys
import time
import json
import csv
import numpy as np
import pandas as pd
from pathlib import Path
from tqdm import tqdm

# GPU 加速支持
try:
    import torch
    TORCH_AVAILABLE = True
except ImportError:
    TORCH_AVAILABLE = False

###############################################################################
# 导入复用函数
###############################################################################
import importlib.util

# 导入碎片匹配函数 (统一使用L1_MC的Numba加速版本)
_l1_mc_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           '..', 'L1_真实数据库匹配', 'L1_MC匹配.py')
_spec_mc = importlib.util.spec_from_file_location("L1_MC", os.path.abspath(_l1_mc_path))
_l1_mc = importlib.util.module_from_spec(_spec_mc)
_spec_mc.loader.exec_module(_l1_mc)

find_matched_fragments = _l1_mc.find_matched_fragments

# 导入L1的库加载函数
_l1_dreams_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                '..', 'L1_真实数据库匹配', 'L1_DreaMS匹配.py')
_spec_dreams = importlib.util.spec_from_file_location("L1_DreaMS", os.path.abspath(_l1_dreams_path))
_l1_dreams = importlib.util.module_from_spec(_spec_dreams)
_spec_dreams.loader.exec_module(_l1_dreams)

load_library = _l1_dreams.load_library

# 导入 Excel 格式化函数
_summary_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             '..', '辅助功能', '各层鉴定结果汇总', '各层鉴定结果汇总.py')
_spec_sum = importlib.util.spec_from_file_location("SUMMARY", os.path.abspath(_summary_path))
_l1_sum = importlib.util.module_from_spec(_spec_sum)
_spec_sum.loader.exec_module(_l1_sum)

format_excel_output = _l1_sum.format_excel_output


###############################################################################
# 轻量 MSP 解析
###############################################################################

def parse_msp_metadata(msp_path):
    """从 MSP 解析元数据 + 峰列表（复用L1缓存机制）"""
    # 直接调用L1的load_library函数
    library = load_library('', msp_path)
    
    # 转换matchms对象为字典格式
    entries = []
    for spec in library:
        # matchms对象中名称可能在compound_name或name字段
        name = spec.get('compound_name') or spec.get('name', '')
        entry = {
            'name': name,
            'peaks': list(spec.peaks) if hasattr(spec, 'peaks') else [],
        }
        # 提取其他元数据
        for key in ['smiles', 'inchikey', 'formula', 'precursor_mz', 
                    'ontology', 'adduct', 'retention_time', 'ion_mode',
                    'source_tool', 'source_db', 'comment']:
            val = spec.get(key)
            if val is not None:
                entry[key] = val
        entries.append(entry)
    
    return entries


###############################################################################
# 核心：纯 Embedding 类似物筛查
###############################################################################

def process_sample_candidates(global_i, sorted_idx, scores_for_result, sample_info, lib_info,
                               fragment_tolerance, min_precursor_ppm, min_matched_peaks):
    """处理单个样品的候选匹配（公共函数）"""
    s_info = sample_info[global_i] if global_i < len(sample_info) else {}
    s_name = s_info.get('name', f'sample_{global_i}')
    s_pmz = s_info.get('precursor_mz', 0)
    s_peaks = s_info.get('peaks', [])

    results = []
    actual_rank = 0
    sample_has_hit = False

    for lib_idx in sorted_idx:
        lib_meta = lib_info[lib_idx] if lib_idx < len(lib_info) else {}
        lib_pmz = lib_meta.get('precursor_mz', 0)
        lib_peaks = lib_meta.get('peaks', [])

        # 计算母离子偏差 (ppm)
        try:
            s_pmz_f = float(s_pmz)
            lib_pmz_f = float(lib_pmz)
            mass_diff = round(s_pmz_f - lib_pmz_f, 4)
            if lib_pmz_f > 0:
                precursor_ppm = abs(s_pmz_f - lib_pmz_f) / lib_pmz_f * 1e6
            else:
                precursor_ppm = float('inf')
        except (ValueError, TypeError):
            mass_diff = None
            precursor_ppm = float('inf')

        # 过滤：母离子偏差 <= min_precursor_ppm 的排除（与L2精确匹配重叠）
        if precursor_ppm <= min_precursor_ppm:
            continue

        # 碎片匹配
        matched_frags = []
        matched_peaks_count = 0
        matched_frags_str = ""
        if s_peaks and lib_peaks:
            matched_frags = find_matched_fragments(
                s_peaks, lib_peaks, fragment_tolerance,
                s_pmz, lib_pmz
            )
            matched_peaks_count = len(matched_frags)
            if matched_frags:
                matched_frags_str = "; ".join(
                    f"Q{m['query_mz']}=L{m['lib_mz']}(Δ{m['mz_diff']})"
                    for m in matched_frags
                )

        # 过滤：匹配碎片数 < min_matched_peaks 的排除
        if matched_peaks_count < min_matched_peaks:
            continue

        actual_rank += 1
        sample_has_hit = True

        # 合并 source_database
        _src_db = lib_meta.get('source_db', '')
        _src_tool = lib_meta.get('source_tool', '')
        if _src_tool and _src_db:
            source_database = f"{_src_tool}/{_src_db}"
        else:
            source_database = _src_db
        
        results.append({
            'query_name': s_name,
            'query_precursor_mz': s_pmz,
            'matched_name': lib_meta.get('name', f'lib_{lib_idx}'),
            'matched_formula': lib_meta.get('formula', ''),
            'matched_smiles': lib_meta.get('smiles', ''),
            'matched_inchikey': lib_meta.get('inchikey', ''),
            'matched_precursor_mz': lib_pmz,
            'matched_ontology': lib_meta.get('ontology', ''),
            'source_database': source_database,
            'analog_rank': actual_rank,
            'cosine_similarity': round(float(scores_for_result[lib_idx]), 4),
            'mass_difference': mass_diff,
            'matched_peaks': matched_peaks_count,
            'matched_fragments': matched_frags_str,
            'adduct': s_info.get('adduct', ''),
            'source_method': 'DreaMS_Analog',
        })

    return results, sample_has_hit


def screen_analogs_by_embedding(sample_embs, lib_embs, sample_info, lib_info,
                                threshold, fragment_tolerance,
                                min_matched_peaks, min_precursor_ppm):
    """纯 Embedding cosine 相似度类似物筛查

    方向：对每个样品，在合并候选库中找所有满足条件的类似物
    过滤条件：
      1. 碎片匹配数 >= min_matched_peaks（确保谱图层面有实际证据）
      2. 母离子偏差 > min_precursor_ppm（排除与L2精确匹配重叠的结果）

    参数:
        sample_embs: 样品 embedding 矩阵 (n_sample, dim)
        lib_embs: 候选库 embedding 矩阵 (n_lib, dim)
        sample_info: 样品元数据列表
        lib_info: 候选库元数据列表
        threshold: cosine 相似度阈值
        fragment_tolerance: 碎片匹配容差 (Da)
        min_matched_peaks: 最少匹配碎片数
        min_precursor_ppm: 母离子最小偏差 (ppm)，≤此值排除

    返回:
        results: 类似物筛查结果列表
        matched_sample_indices: 有命中结果的样品索引集合
    """
    n_sample = len(sample_embs)
    n_lib = len(lib_embs)

    # 检测GPU可用性（强制使用单GPU，多GPU分片存在跨设备计算bug）
    use_gpu = TORCH_AVAILABLE and torch.cuda.is_available()

    # 归一化 embedding（计算 cosine 相似度）
    if use_gpu:
        # GPU路径：使用PyTorch（单GPU cuda:0）
        sample_tensor = torch.tensor(sample_embs, dtype=torch.float32, device='cuda:0')
        lib_tensor = torch.tensor(lib_embs, dtype=torch.float32, device='cuda:0')
        
        sample_norms = torch.norm(sample_tensor, dim=1, keepdim=True)
        lib_norms = torch.norm(lib_tensor, dim=1, keepdim=True)
        sample_norms = torch.where(sample_norms == 0, torch.ones_like(sample_norms), sample_norms)
        lib_norms = torch.where(lib_norms == 0, torch.ones_like(lib_norms), lib_norms)
        
        sample_normed = sample_tensor / sample_norms
        lib_normed = lib_tensor / lib_norms
    else:
        # CPU路径：使用numpy
        sample_norms = np.linalg.norm(sample_embs, axis=1, keepdims=True)
        lib_norms = np.linalg.norm(lib_embs, axis=1, keepdims=True)
        sample_norms[sample_norms == 0] = 1.0
        lib_norms[lib_norms == 0] = 1.0
        sample_normed = sample_embs / sample_norms
        lib_normed = lib_embs / lib_norms

    results = []
    matched_sample_indices = set()

    # 分批计算避免 OOM
    batch_size = 500 if use_gpu else 100  # GPU可用更大批次
    
    # 使用样品级别的进度条
    pbar = tqdm(total=n_sample, desc="  类似物筛查", ncols=100)
    
    for batch_start in range(0, n_sample, batch_size):
        batch_end = min(batch_start + batch_size, n_sample)
        batch_sample = sample_normed[batch_start:batch_end]

        # 计算当前批次与全库的相似度矩阵 (batch, n_lib)
        if use_gpu:
            # 单GPU（cuda:0）
            sim_matrix = torch.mm(batch_sample, lib_normed.t())
            
            for local_i in range(batch_end - batch_start):
                global_i = batch_start + local_i
                scores = sim_matrix[local_i]
                above_mask = scores >= threshold
                above = torch.where(above_mask)[0].cpu().numpy()
                if len(above) == 0:
                    pbar.update(1)
                    continue
                scores_np = scores.cpu().numpy()
                sorted_idx = above[np.argsort(scores_np[above])[::-1]]
                
                sample_results, has_hit = process_sample_candidates(
                    global_i, sorted_idx, scores_np, sample_info, lib_info,
                    fragment_tolerance, min_precursor_ppm, min_matched_peaks
                )
                results.extend(sample_results)
                if has_hit:
                    matched_sample_indices.add(global_i)
                
                pbar.update(1)
        else:
            # CPU路径
            sim_matrix = batch_sample @ lib_normed.T
            
            for local_i in range(batch_end - batch_start):
                global_i = batch_start + local_i
                scores = sim_matrix[local_i]
                above = np.where(scores >= threshold)[0]
                if len(above) == 0:
                    pbar.update(1)
                    continue
                
                sorted_idx = above[np.argsort(scores[above])[::-1]]
                
                sample_results, has_hit = process_sample_candidates(
                    global_i, sorted_idx, scores, sample_info, lib_info,
                    fragment_tolerance, min_precursor_ppm, min_matched_peaks
                )
                results.extend(sample_results)
                if has_hit:
                    matched_sample_indices.add(global_i)
                
                pbar.update(1)
    
    pbar.close()

    return results, matched_sample_indices


###############################################################################
# 列名翻译 + Excel
###############################################################################

def _analog_column_rename():
    """类似物筛查列名：英文 → 中文映射"""
    return {
        'query_name': '样品化合物',
        'query_precursor_mz': '样品母离子(m/z)',
        'matched_name': '匹配化合物',
        'matched_formula': '匹配分子式',
        'matched_smiles': '匹配SMILES',
        'matched_inchikey': '匹配InChIKey',
        'matched_precursor_mz': '匹配母离子(m/z)',
        'matched_ontology': '分类',
        'source_database': '来源数据库',
        'analog_rank': '类似物排名',
        'cosine_similarity': 'DreaMS余弦相似度',
        'mass_difference': '质量差(Da)',
        'matched_peaks': '匹配碎片数',
        'matched_fragments': '匹配碎片详情',
        'adduct': '加合物类型',
        'source_method': '匹配方法',
    }


def _save_analog_excel(df, excel_path):
    """将英文列名 DataFrame 转为中文列名 Excel
    
    对于超大结果集（>100万行），自动分多个sheet存储
    """
    MAX_ROWS_PER_SHEET = 1000000  # Excel单sheet最大行数
    
    df_display = df.rename(columns=_analog_column_rename())
    n_rows = len(df_display)
    
    if n_rows <= MAX_ROWS_PER_SHEET:
        format_excel_output(df_display, excel_path, wrap_columns="匹配化合物,分类,匹配碎片详情")
    else:
        # 分多个sheet
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        n_sheets = (n_rows + MAX_ROWS_PER_SHEET - 1) // MAX_ROWS_PER_SHEET
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            for i in range(n_sheets):
                start_idx = i * MAX_ROWS_PER_SHEET
                end_idx = min((i + 1) * MAX_ROWS_PER_SHEET, n_rows)
                df_chunk = df_display.iloc[start_idx:end_idx]
                sheet_name = f'结果_{i+1}'
                df_chunk.to_excel(writer, index=False, sheet_name=sheet_name)
                
                ws = writer.sheets[sheet_name]
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                for col_idx, cell in enumerate(ws[1], start=1):
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                wrap_cols = ['匹配化合物', '分类', '匹配碎片详情']
                for col_idx, col_name in enumerate(df_chunk.columns, start=1):
                    if col_name in wrap_cols:
                        for row_idx in range(2, ws.max_row + 1):
                            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True)


def regenerate_analog_excel(csv_path, excel_path):
    """从已处理的 CSV（ontology + 翻译后）重新生成中文列名 Excel

    用途：总控脚本在 ontology 获取 + 翻译完成后调用
    """
    # 检查文件是否存在且非空
    if not os.path.exists(csv_path):
        print(f"  [警告] CSV文件不存在: {csv_path}")
        return
    
    # 检查文件大小
    if os.path.getsize(csv_path) == 0:
        print(f"  [警告] CSV文件为空: {csv_path}")
        return
    
    try:
        df = pd.read_csv(csv_path, encoding='utf-8')
        if df.empty:
            print(f"  [警告] CSV无数据: {csv_path}")
            return
        _save_analog_excel(df, excel_path)
    except pd.errors.EmptyDataError:
        print(f"  [警告] CSV格式错误(空文件): {csv_path}")
        return


###############################################################################
# MSP 导出（供 L4a/L4b 使用）
###############################################################################

def export_msp_subset(sample_msp_entries, indices, output_path):
    """将指定索引的 MSP 条目导出为新的 MSP 文件

    参数:
        sample_msp_entries: 完整的样品 MSP 条目列表
        indices: 要导出的条目索引集合
        output_path: 输出 MSP 文件路径
    """
    count = 0
    with open(output_path, 'w', encoding='utf-8') as f:
        for idx in sorted(indices):
            if idx >= len(sample_msp_entries):
                continue
            entry = sample_msp_entries[idx]
            f.write(f"NAME: {entry.get('name', 'Unknown')}\n")
            if 'precursor_mz' in entry:
                f.write(f"PRECURSORMZ: {entry['precursor_mz']}\n")
            if 'adduct' in entry:
                f.write(f"PRECURSORTYPE: {entry['adduct']}\n")
            if 'ion_mode' in entry:
                f.write(f"IONMODE: {entry['ion_mode']}\n")
            if 'formula' in entry:
                f.write(f"FORMULA: {entry['formula']}\n")
            if 'smiles' in entry:
                f.write(f"SMILES: {entry['smiles']}\n")
            if 'inchikey' in entry:
                f.write(f"INCHIKEY: {entry['inchikey']}\n")
            if 'retention_time' in entry:
                f.write(f"RETENTIONTIME: {entry['retention_time']}\n")
            if 'ccs' in entry:
                f.write(f"CCS: {entry['ccs']}\n")
            if 'comment' in entry:
                f.write(f"COMMENT: {entry['comment']}\n")
            peaks = entry.get('peaks', [])
            if peaks:
                f.write(f"Num Peaks: {len(peaks)}\n")
                for mz, intensity in peaks:
                    f.write(f"{mz}\t{intensity}\n")
            f.write("\n")
            count += 1
    return count


###############################################################################
# 主函数
###############################################################################

def main():
    # ---- 读取 & 验证环境变量 ----
    OUTPUT_DIR        = os.environ.get('L3_OUTPUT_DIR')
    SAMPLE_EMB        = os.environ.get('L3_SAMPLE_EMB')
    SAMPLE_MSP        = os.environ.get('L3_SAMPLE_MSP')  # L2_unidentified.msp
    SIM_THRESHOLD     = os.environ.get('L3_SIM_THRESHOLD')
    FRAG_TOLERANCE    = os.environ.get('L3_FRAGMENT_TOLERANCE', '0.05')
    MIN_MATCHED_PEAKS = os.environ.get('L3_MIN_MATCHED_PEAKS')
    MIN_PRECURSOR_PPM = os.environ.get('L3_MIN_PRECURSOR_PPM')

    # 候选库（真实库 + 模拟库）
    REAL_DB_LIBRARIES = os.environ.get('L3_REAL_DB_LIBRARIES')  # JSON: {name: {msp:..., emb:...}}
    SIMULATED_MSP     = os.environ.get('L3_SIMULATED_MSP', '')
    SIMULATED_EMB     = os.environ.get('L3_SIMULATED_EMB', '')

    _missing = []
    for var_name, var_val in [
        ('L3_OUTPUT_DIR', OUTPUT_DIR), ('L3_SAMPLE_EMB', SAMPLE_EMB),
        ('L3_SAMPLE_MSP', SAMPLE_MSP), ('L3_SIM_THRESHOLD', SIM_THRESHOLD),
        ('L3_REAL_DB_LIBRARIES', REAL_DB_LIBRARIES),
        ('L3_MIN_MATCHED_PEAKS', MIN_MATCHED_PEAKS),
        ('L3_MIN_PRECURSOR_PPM', MIN_PRECURSOR_PPM),
    ]:
        if not var_val:
            _missing.append(var_name)
    if _missing:
        raise ValueError(f"错误：以下环境变量未设置：{', '.join(_missing)}")

    SIM_THRESHOLD  = float(SIM_THRESHOLD)
    FRAG_TOLERANCE = float(FRAG_TOLERANCE)
    MIN_MATCHED_PEAKS = int(MIN_MATCHED_PEAKS)
    MIN_PRECURSOR_PPM = float(MIN_PRECURSOR_PPM)

    start_time = time.time()

    print("=" * 70)
    print("L3: DreaMS 类似物筛查（纯 Embedding 语义匹配）")
    print("=" * 70)
    print(f"  样品 Embedding: {SAMPLE_EMB}")
    print(f"  样品 MSP: {SAMPLE_MSP}")
    print(f"  相似度阈值: {SIM_THRESHOLD}")
    print(f"  碎片容差: {FRAG_TOLERANCE} Da")
    print(f"  最少匹配碎片数: {MIN_MATCHED_PEAKS}（低于此值过滤）")
    print(f"  母离子最小偏差: {MIN_PRECURSOR_PPM} ppm（小于此值过滤，避免与L2重叠）")
    print(f"  输出目录: {OUTPUT_DIR}")
    print("=" * 70)

    output_dir = Path(OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)

    # ======== 1. 加载样品 embedding + MSP ========
    print("\n[1/5] 加载样品数据...")

    if not os.path.exists(SAMPLE_EMB):
        print(f"[ERROR] 样品 embedding 不存在: {SAMPLE_EMB}")
        return False
    if not os.path.exists(SAMPLE_MSP):
        print(f"[ERROR] 样品 MSP 不存在: {SAMPLE_MSP}")
        return False

    sample_data = np.load(SAMPLE_EMB, allow_pickle=True)
    sample_embs_full = sample_data['embeddings'].astype(np.float32)
    sample_msp_entries = parse_msp_metadata(SAMPLE_MSP)

    # 匹配 embedding 和 MSP（通过名称）
    sample_info = []
    sample_emb_indices = []

    if sample_msp_entries and 'names' in sample_data:
        emb_names = list(sample_data['names'])
        emb_pmzs = list(sample_data.get('precursor_mzs', [0.0] * len(emb_names)))
        name_to_idx = {name: i for i, name in enumerate(emb_names)}

        for spec in sample_msp_entries:
            name = spec.get('name', '')
            if name in name_to_idx:
                idx = name_to_idx[name]
                sample_info.append({
                    'name': emb_names[idx],
                    'precursor_mz': float(emb_pmzs[idx]) if idx < len(emb_pmzs) else spec.get('precursor_mz', 0),
                    'peaks': spec.get('peaks', []),
                    'adduct': spec.get('adduct', ''),
                })
                sample_emb_indices.append(idx)

        if sample_emb_indices:
            sample_embs = sample_embs_full[sample_emb_indices]
        else:
            sample_embs = sample_embs_full
            sample_info = sample_msp_entries
        print(f"  样品: MSP {len(sample_msp_entries)} 条 → Embedding 匹配 {len(sample_info)} 条")
    else:
        sample_embs = sample_embs_full
        sample_info = sample_msp_entries
        print(f"  样品: {sample_embs.shape[0]} 条（无名称匹配，直接对齐）")

    n_sample = len(sample_embs)

    # ======== 2. 加载候选库 embedding + MSP（真实库 + 模拟库合并） ========
    print("\n[2/5] 加载候选库（真实库 + 模拟库）...")

    all_lib_embs = []
    all_lib_info = []

    # 2a: 真实库
    real_db_config = json.loads(REAL_DB_LIBRARIES)
    print(f"[诊断] 真实数据库配置: {list(real_db_config.keys())}")

    for db_name, db_paths in real_db_config.items():
        emb_path = db_paths.get('emb', '')
        msp_path = db_paths.get('msp', '')

        print(f"[诊断] 处理数据库 {db_name}:")
        print(f"  EMB路径: {emb_path}")
        print(f"  EMB存在: {os.path.exists(emb_path) if emb_path else False}")

        if not emb_path or not os.path.exists(emb_path):
            print(f"  ❌ 跳过此数据库")
            continue

        lib_data = np.load(emb_path, allow_pickle=True)
        lib_embs = lib_data['embeddings'].astype(np.float32)

        # 元数据优先从 NPZ 读取
        lib_entries = []
        if 'names' in lib_data:
            names = list(lib_data['names'])
            smiles_arr = list(lib_data.get('smiles', [''] * len(names)))
            inchikey_arr = list(lib_data.get('inchikeys', [''] * len(names)))
            formula_arr = list(lib_data.get('formulas', [''] * len(names)))
            pmz_arr = list(lib_data.get('precursor_mzs', [0.0] * len(names)))
            ontology_arr = list(lib_data.get('ontologies', [''] * len(names)))

            for j in range(len(names)):
                lib_entries.append({
                    'name': names[j],
                    'smiles': smiles_arr[j] if j < len(smiles_arr) else '',
                    'inchikey': inchikey_arr[j] if j < len(inchikey_arr) else '',
                    'formula': formula_arr[j] if j < len(formula_arr) else '',
                    'precursor_mz': float(pmz_arr[j]) if j < len(pmz_arr) else 0.0,
                    'ontology': ontology_arr[j] if j < len(ontology_arr) else '',
                    'source_db': db_name,
                    'peaks': [],  # NPZ 无峰列表
                })
        else:
            lib_entries = [{'source_db': db_name, 'peaks': []}] * len(lib_embs)

        # 加载 MSP 峰列表（碎片匹配用）
        if msp_path and os.path.exists(msp_path):
            msp_entries = parse_msp_metadata(msp_path)
            msp_name_map = {entry.get('name', ''): entry.get('peaks', []) for entry in msp_entries}
            
            # 诊断: 检查名称匹配
            print(f"    [诊断] MSP条目数: {len(msp_entries)}, NPZ条目数: {len(lib_entries)}")
            if msp_entries and lib_entries:
                sample_msp_name = msp_entries[0].get('name', '')[:50]
                sample_npz_name = lib_entries[0].get('name', '')[:50]
                print(f"    [诊断] MSP示例名称: {sample_msp_name}")
                print(f"    [诊断] NPZ示例名称: {sample_npz_name}")
            
            matched_count = 0
            for entry in lib_entries:
                name = entry.get('name', '')
                if name in msp_name_map:
                    entry['peaks'] = msp_name_map[name]
                    if entry['peaks']:
                        matched_count += 1
            print(f"    [峰列表] {db_name}: {matched_count}/{len(lib_entries)} 条有峰数据")
        else:
            print(f"    [峰列表] {db_name}: MSP文件不存在,无峰数据")

        all_lib_embs.append(lib_embs)
        all_lib_info.extend(lib_entries)

    # 2b: 模拟库
    if SIMULATED_EMB and os.path.exists(SIMULATED_EMB):
        sim_data = np.load(SIMULATED_EMB, allow_pickle=True)
        sim_embs = sim_data['embeddings'].astype(np.float32)

        sim_entries = []
        if 'names' in sim_data:
            names = list(sim_data['names'])
            smiles_arr = list(sim_data.get('smiles', [''] * len(names)))
            inchikey_arr = list(sim_data.get('inchikeys', [''] * len(names)))
            formula_arr = list(sim_data.get('formulas', [''] * len(names)))
            pmz_arr = list(sim_data.get('precursor_mzs', [0.0] * len(names)))
            ontology_arr = list(sim_data.get('ontologies', [''] * len(names)))

            for j in range(len(names)):
                sim_entries.append({
                    'name': names[j],
                    'smiles': smiles_arr[j] if j < len(smiles_arr) else '',
                    'inchikey': inchikey_arr[j] if j < len(inchikey_arr) else '',
                    'formula': formula_arr[j] if j < len(formula_arr) else '',
                    'precursor_mz': float(pmz_arr[j]) if j < len(pmz_arr) else 0.0,
                    'ontology': ontology_arr[j] if j < len(ontology_arr) else '',
                    'source_db': 'simulated',
                    'peaks': [],
                })
        else:
            sim_entries = [{'source_db': 'simulated', 'peaks': []}] * len(sim_embs)

        # 加载模拟库 MSP 峰列表 + 提取 source_tool/source_database
        if SIMULATED_MSP and os.path.exists(SIMULATED_MSP):
            msp_entries = parse_msp_metadata(SIMULATED_MSP)
            msp_name_map = {}
            msp_source_map = {}  # name -> (source_tool, source_database)
            for entry in msp_entries:
                name = entry.get('name', '')
                msp_name_map[name] = entry.get('peaks', [])
                # 从 comment 提取 source 信息
                comment = entry.get('comment', '') or ''
                source_tool = ''
                source_db = ''
                if comment and 'Source_tool=' in comment:
                    for part in comment.split(';'):
                        part = part.strip()
                        if part.startswith('Source_tool='):
                            source_tool = part.split('=', 1)[1].strip()
                        elif part.startswith('Source_database='):
                            source_db = part.split('=', 1)[1].strip()
                if source_tool or source_db:
                    msp_source_map[name] = (source_tool, source_db)
            for entry in sim_entries:
                name = entry.get('name', '')
                if name in msp_name_map:
                    entry['peaks'] = msp_name_map[name]
                if name in msp_source_map:
                    source_tool, source_db = msp_source_map[name]
                    if source_tool:
                        entry['source_tool'] = source_tool
                    if source_db:
                        entry['source_db'] = source_db

        all_lib_embs.append(sim_embs)
        all_lib_info.extend(sim_entries)

    if not all_lib_embs:
        print("[ERROR] 无可用候选库 embedding")
        return False

    # 合并所有候选库
    merged_lib_embs = np.vstack(all_lib_embs)

    # ======== 3. 类似物筛查 ========
    print(f"\n[3/5] 类似物筛查（阈值 ≥ {SIM_THRESHOLD}）...")

    analogs, matched_indices = screen_analogs_by_embedding(
        sample_embs, merged_lib_embs, sample_info, all_lib_info,
        threshold=SIM_THRESHOLD,
        fragment_tolerance=FRAG_TOLERANCE,
        min_matched_peaks=MIN_MATCHED_PEAKS,
        min_precursor_ppm=MIN_PRECURSOR_PPM
    )

    # ======== 4. 导出结果 ========
    print(f"\n[4/5] 导出结果...")

    if analogs:
        df = pd.DataFrame(analogs)
        df = df.sort_values(['query_name', 'analog_rank'])
        csv_path = str(output_dir / 'L3_results.csv')
        df.to_csv(csv_path, index=False, encoding='utf-8')
        
        excel_path = str(output_dir / 'L3_results.xlsx')
        _save_analog_excel(df, excel_path)
        
        unique_samples = len(set(a['query_name'] for a in analogs))
        print(f"  CSV: {len(df)} 条, {unique_samples} 个样品命中")
    else:
        print("  无命中结果")
        csv_path = str(output_dir / 'L3_results.csv')
        pd.DataFrame().to_csv(csv_path, index=False, encoding='utf-8')

    # ======== 5. 生成 L4a/L4b 输入 MSP ========
    print(f"\n[5/5] 生成 L4 输入文件...")

    analog_msp_path = str(output_dir / 'L3_analog.msp')
    n_analog = export_msp_subset(sample_msp_entries, matched_indices, analog_msp_path)

    all_indices = set(range(len(sample_msp_entries)))
    unmatched_indices = all_indices - matched_indices
    unid_msp_path = str(output_dir / 'L3_remaining.msp')
    n_unid = export_msp_subset(sample_msp_entries, unmatched_indices, unid_msp_path)

    elapsed = time.time() - start_time

    print("\n" + "=" * 70)
    print(f"L3 DreaMS 类似物筛查完成 | 样品: {n_sample} | 候选库: {merged_lib_embs.shape[0]} | 命中: {len(matched_indices)} | 结果: {len(analogs)} 条 | 耗时: {elapsed:.1f}s")
    print(f"  → L4a: {n_analog} 条（类似物） | L4b: {n_unid} 条（未知物）")
    print("=" * 70)

    return True


if __name__ == "__main__":
    # 支持 --regenerate_excel 模式：辅助功能处理后重新生成 Excel
    if len(sys.argv) >= 3 and sys.argv[1] == '--regenerate_excel':
        csv_path = sys.argv[2]
        excel_path = sys.argv[3] if len(sys.argv) >= 4 else csv_path.replace('.csv', '.xlsx')
        regenerate_analog_excel(csv_path, excel_path)
        sys.exit(0)
    else:
        success = main()
        sys.exit(0 if success else 1)
