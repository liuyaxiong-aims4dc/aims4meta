#!/usr/bin/env python3
"""
多层鉴定结果汇总工具

统一处理：
- 各层（L1/L2/L3/L4）结果汇总
- 最终结果汇总（L1+L2+L3+L4a+L4b）
- 格式化Excel输出
"""

import argparse
import os
import sys
import pandas as pd
import numpy as np
from pathlib import Path
from typing import List, Optional, Dict


def parse_arguments():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(description="多层鉴定结果汇总")
    parser.add_argument("--mode", required=True, 
                       choices=['L1', 'L2', 'L3', 'L4', 'final'],
                       help="汇总模式: L1/L2/L3(DreaMS类似物)/L4(SIRIUS从头鉴定)/final")
    parser.add_argument("--input", default=None,
                       help="输入CSV文件路径（L1/L2/L3模式）或结果目录（final模式）")
    parser.add_argument("--mc_input", default=None, help="MC匹配结果CSV路径（L1/L2模式，与--dreams_input搭配使用）")
    parser.add_argument("--dreams_input", default=None, help="DreaMS匹配结果CSV路径（L1/L2模式，与--mc_input搭配使用）")
    parser.add_argument("--output", required=True, help="输出文件路径")
    parser.add_argument("--skip_excel", action="store_true", help="跳过Excel生成（仅输出CSV，用于中间整合步骤）")
    parser.add_argument("--l1_results", help="L1结果CSV路径（final模式）")
    parser.add_argument("--l2_results", help="L2结果CSV路径（final模式）")
    parser.add_argument("--l3_results", help="L3结果CSV路径（final模式）")
    parser.add_argument("--l4a_results", help="L4a类似物鉴定结果CSV路径（final模式）")
    parser.add_argument("--l4b_results", help="L4b未知物鉴定结果CSV路径（final模式）")
    # 生成未鉴定MSP文件的参数（已废弃，各层独立分析原始数据，不再生成unidentified MSP）
    parser.add_argument("--sample_msp", help="[已废弃] 原始样品MSP文件路径")
    parser.add_argument("--sample_csv", help="[已废弃] 原始样品CSV文件路径")
    parser.add_argument("--additional_identified_csv", help="[已废弃] 额外已鉴定结果CSV")
    return parser.parse_args()


def get_column_order(level: str) -> List[str]:
    """
    获取各层的标准列顺序
    
    参数:
        level: 鉴定层级 (L1/L2/L3)
    
    返回:
        标准列顺序列表
    """
    # 基础列顺序（所有层级通用）
    base_order = [
        # 1. 查询化合物信息
        'query_name',
        # 2. 匹配化合物基本信息
        'matched_name', 'matched_smiles', 'matched_inchikey', 'matched_formula',
        # 3. 匹配质量指标
        'cosine_score', 'matched_peaks', 'matched_fragments',
        # 4. 质谱信息
        'precursor_mz', 'library_precursor_mz', 'precursor_ppm_diff', 'adduct',
        # 5. Ontology信息
        'matched_ontology',
        # 6. 来源信息
        'source_method', 'source_database', 'comprehensive_score',
        # 7. 原始样本信息（从原始CSV关联）
        'Retention time (min)', 'CCS (angstrom^2)', 'Maximum Abundance',
        # 8. CCS信息（全部放在一起）
        'library_ccs', 'predicted_ccs', 'prediction_source',
        'library_ccs_deviation_pct', 'predicted_ccs_deviation_pct',
        # 9. 同位素信息（放在最后）
        'Isotope Distribution', 'isotope_similarity',
    ]
    
    # L2特有列
    l2_specific = ['simulated_library', 'source_tool']
    
    # L4特有列（SIRIUS从头鉴定）
    l4_specific = ['formula', 'structure_confidence', 'sirius_score', 'zodiac_score']
    
    if level == 'L1':
        return base_order
    elif level == 'L2':
        # L2在来源信息后添加特有列
        insert_pos = base_order.index('comprehensive_score') + 1
        return base_order[:insert_pos] + l2_specific + base_order[insert_pos:]
    elif level == 'L3':
        # L3 DreaMS类似物筛查：与L1/L2类似，但无严格匹配限制
        return base_order
    elif level == 'L4':
        # L4 SIRIUS从头鉴定：列顺序不同（没有library_precursor_mz等）
        return [
            'query_name',
            'matched_name', 'formula', 'matched_inchikey',
            'sirius_score', 'zodiac_score', 'structure_confidence',
            'precursor_mz', 'adduct',
            'Retention time (min)', 'CCS (angstrom^2)',
            'Isotope Distribution', 'isotope_similarity',
        ]
    
    return base_order


def get_column_rename_map(level: str) -> dict:
    """
    获取各层的列名中文映射
    
    参数:
        level: 鉴定层级 (L1/L2/L3)
    
    返回:
        {英文列名: 中文列名} 字典
    """
    # 通用列名映射
    base_map = {
        # 1. 查询化合物信息
        'query_name': '样品化合物',
        # 2. 匹配化合物基本信息
        'matched_name': '匹配化合物',
        'matched_smiles': '匹配SMILES',
        'matched_inchikey': '匹配InChIKey',
        'matched_formula': '匹配分子式',
        # 3. 匹配质量指标
        'cosine_score': '余弦相似度',
        'matched_peaks': '匹配碎片数',
        'matched_fragments': '匹配碎片详情',
        # 4. 质谱信息
        'precursor_mz': '样品母离子(m/z)',
        'library_precursor_mz': '库母离子(m/z)',
        'precursor_ppm_diff': '母离子偏差(ppm)',
        'adduct': '加合物类型',
        # 5. Ontology信息
        'matched_ontology': '分类',
        # 6. 来源信息
        'source_method': '匹配方法',
        'source_database': '来源数据库',
        'comprehensive_score': '综合得分',
        'rank': '排名',
        # 7. 原始样本信息
        'Retention time (min)': '保留时间(min)',
        'CCS (angstrom^2)': '实验CCS(Å²)',
        'Maximum Abundance': '最大丰度',
        # 8. CCS信息
        'library_ccs': '库CCS(Å²)',
        'predicted_ccs': '预测CCS(Å²)',
        'prediction_source': 'CCS预测来源',
        'library_ccs_deviation_pct': '库CCS偏差(%)',
        'predicted_ccs_deviation_pct': '预测CCS偏差(%)',
        # 9. 同位素信息
        'Isotope Distribution': '同位素分布',
        'isotope_similarity': '同位素相似度',
        # 10. 其他
        'matched_name_cn': '匹配化合物(中文)',
    }
    
    # L2特有列
    l2_map = {
        'simulated_library': '模拟库',
        'source_tool': '预测工具',
    }
    
    if level == 'L2':
        base_map.update(l2_map)
    
    return base_map


def get_columns_to_drop(level: str) -> List[str]:
    """
    获取各层需要删除的列
    
    参数:
        level: 鉴定层级 (L1/L2/L3)
    
    返回:
        需要删除的列名列表
    """
    # 所有层级都不需要的列
    common_drop = [
        'sample_charge',
        # 原始CSV中不需要保留的列
        'Neutral mass (Da)', 'm/z', 'Charge',
        'Chromatographic peak width (min)', 'Identifications',
        'Minimum CV%',
        # 原始CSV中的丰度列（样本特定名称）
        '20260203-LIHUA-HDMSE-POS', '20260203-LIHUA-HDMSE-POS.1',
        # 原始CSV中的已接受信息（冗余）
        'Accepted Compound ID', 'Accepted Description',
        # 原始CSV中的加合物/分子式/得分（我们有自己的matched_版本）
        'Adducts', 'Formula', 'Score', 'Fragmentation Score',
        'Mass Error (ppm)', 'Isotope Similarity',
        'Retention Time Error (mins)', 'dCCS (angstrom^2)', 'Compound Link',
    ]
    
    if level == 'L1':
        return common_drop
    elif level == 'L2':
        return common_drop
    elif level == 'L3':
        # L3 DreaMS类似物筛查
        return common_drop
    elif level == 'L4':
        # L4 SIRIUS从头鉴定可能没有这些列
        return common_drop + ['library_ccs', 'library_precursor_mz', 'precursor_ppm_diff']
    
    return common_drop


def get_wrap_columns(level: str) -> str:
    """
    获取需要自动换行的列
    
    参数:
        level: 鉴定层级 (L1/L2/L3)
    
    返回:
        逗号分隔的列名字符串
    """
    if level == 'L4':
        return "matched_name,formula"
    else:
        return "matched_name,matched_ontology,matched_fragments"


def merge_mc_dreams(mc_csv: str, dreams_csv: str) -> Optional[pd.DataFrame]:
    """
    合并 MC 和 DreaMS 匹配结果（并集策略）
    
    参数:
        mc_csv: MC匹配结果CSV路径
        dreams_csv: DreaMS匹配结果CSV路径
    
    返回:
        合并后的DataFrame，失败返回None
    """
    if mc_csv and os.path.exists(mc_csv):
        mc_df = pd.read_csv(mc_csv)
        mc_count = len(mc_df)
    else:
        mc_df = pd.DataFrame()
        mc_count = 0
    
    if dreams_csv and os.path.exists(dreams_csv):
        dreams_df = pd.read_csv(dreams_csv)
        dreams_count = len(dreams_df)
    else:
        dreams_df = pd.DataFrame()
        dreams_count = 0
    
    if mc_df.empty and dreams_df.empty:
        return None
    
    # 合并
    if mc_df.empty:
        combined = dreams_df.copy()
    elif dreams_df.empty:
        combined = mc_df.copy()
    else:
        combined = pd.concat([mc_df, dreams_df], ignore_index=True)
    
    # 按 comprehensive_score 排序（如存在）
    if 'comprehensive_score' in combined.columns:
        combined = combined.sort_values('comprehensive_score', ascending=False)
    elif 'cosine_score' in combined.columns:
        combined = combined.sort_values('cosine_score', ascending=False)
    
    # 按 (query_name, matched_smiles) 去重（保留最高分）
    if 'query_name' in combined.columns and 'matched_smiles' in combined.columns:
        combined = combined.drop_duplicates(subset=['query_name', 'matched_smiles'], keep='first')
    elif 'query_name' in combined.columns:
        combined = combined.drop_duplicates(subset=['query_name'], keep='first')
    
    combined = combined.reset_index(drop=True)
    
    # 按 source_method 分组重新计算 rank
    if 'query_name' in combined.columns and 'source_method' in combined.columns and 'rank' in combined.columns:
        if 'cosine_score' in combined.columns:
            combined = combined.sort_values(['query_name', 'source_method', 'cosine_score'], ascending=[True, True, False])
        combined['rank'] = combined.groupby(['query_name', 'source_method']).cumcount() + 1
        combined = combined.reset_index(drop=True)
    
    print(f"  MC: {mc_count} + DreaMS: {dreams_count} → 整合: {len(combined)} 条")
    return combined


def process_level_results(input_csv: str, output_csv: str, level: str,
                         sample_msp: str = None, sample_csv: str = None,
                         mc_input: str = None, dreams_input: str = None,
                         additional_identified_csv: str = None,
                         skip_excel: bool = False) -> bool:
    """
    处理指定层级的鉴定结果
    
    参数:
        input_csv: 输入CSV文件路径（已合并的单文件模式）
        output_csv: 输出CSV文件路径
        level: 鉴定层级 (L1/L2/L3)
        mc_input: MC匹配结果CSV路径（MC+DreaMS整合模式）
        dreams_input: DreaMS匹配结果CSV路径（MC+DreaMS整合模式）
        sample_msp/sample_csv/additional_identified_csv: 已废弃，不再使用
    
    返回:
        是否成功
    """
    print(f"\n{'='*60}")
    print(f"{level} 结果汇总")
    print(f"{'='*60}")
    
    # 1. 加载数据
    if mc_input or dreams_input:
        df = merge_mc_dreams(mc_input, dreams_input)
        if df is None:
            return False
    elif input_csv:
        try:
            df = pd.read_csv(input_csv)
            print(f"  加载: {len(df)} 条记录")
        except Exception as e:
            print(f"[ERROR] 加载失败: {e}")
            return False
    else:
        print("[ERROR] 未提供输入")
        return False
    
    # 2. 优化列顺序
    column_order = get_column_order(level)
    columns_to_drop = get_columns_to_drop(level)
    wrap_columns = get_wrap_columns(level)
    
    existing_columns = [col for col in column_order if col in df.columns]
    other_columns = [col for col in df.columns if col not in column_order]
    other_columns = [col for col in other_columns if col not in columns_to_drop]
    
    final_columns = existing_columns + other_columns
    df = df[final_columns]
    
    # 3. 保存CSV
    try:
        df.to_csv(output_csv, index=False, encoding='utf-8')
    except Exception as e:
        print(f"[ERROR] CSV保存失败: {e}")
        return False
    
    # 4. 输出Excel格式
    if not skip_excel:
        rename_map = get_column_rename_map(level)
        df_cn = df.rename(columns=rename_map)
        wrap_columns_cn = [rename_map.get(col, col) for col in wrap_columns.split(',') if col]
        wrap_columns_str = ','.join(wrap_columns_cn)
        
        excel_script = Path(__file__).parent.parent / "Excel格式化输出" / "Excel格式化输出.py"
        if excel_script.exists():
            import subprocess
            temp_csv = output_csv.replace('.csv', '_cn_temp.csv')
            df_cn.to_csv(temp_csv, index=False, encoding='utf-8')
            excel_cmd = [sys.executable, str(excel_script), "--input", temp_csv, "--wrap_columns", wrap_columns_str]
            try:
                subprocess.run(excel_cmd, check=True, capture_output=True)
                temp_xlsx = temp_csv.replace('.csv', '.xlsx')
                final_xlsx = output_csv.replace('.csv', '.xlsx')
                if os.path.exists(temp_xlsx):
                    if os.path.exists(final_xlsx):
                        os.remove(final_xlsx)
                    os.rename(temp_xlsx, final_xlsx)
                if os.path.exists(temp_csv):
                    os.remove(temp_csv)
            except Exception:
                if os.path.exists(temp_csv):
                    os.remove(temp_csv)
        else:
            format_excel_output(df_cn, output_csv.replace('.csv', '.xlsx'), wrap_columns=wrap_columns_str)
    
    # 5. 删除中间CSV文件
    output_dir = os.path.dirname(output_csv)
    for csv_file in [os.path.join(output_dir, f"{level}_MC_results.csv"),
                     os.path.join(output_dir, f"{level}_DreaMS_results.csv")]:
        if os.path.exists(csv_file):
            try:
                os.remove(csv_file)
            except Exception:
                pass
    
    return True


def format_excel_output(df: pd.DataFrame, output_path: str, 
                       wrap_columns: str = ""):
    """
    将DataFrame格式化为Excel文件（内置功能）
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return
    
    # 确保输出路径是.xlsx
    if not output_path.endswith('.xlsx'):
        output_path = output_path.replace('.csv', '.xlsx')
        if not output_path.endswith('.xlsx'):
            output_path += '.xlsx'
    
    # 解析需要换行的列
    wrap_cols = [c.strip() for c in wrap_columns.split(',') if c.strip()]
    wrap_col_indices = []
    for col_name in wrap_cols:
        if col_name in df.columns:
            col_idx = df.columns.get_loc(col_name) + 1
            wrap_col_indices.append(col_idx)
    
    # 按样品化合物排序（支持英文query_name和中文样品化合物）
    sort_col = None
    if 'query_name' in df.columns:
        sort_col = 'query_name'
    elif '样品化合物' in df.columns:
        sort_col = '样品化合物'
    
    if sort_col:
        # 检查可用的辅助排序列（支持中英文列名）
        rank_col = 'rank' if 'rank' in df.columns else ('排名' if '排名' in df.columns else None)
        method_col = 'source_method' if 'source_method' in df.columns else ('匹配方法' if '匹配方法' in df.columns else None)
        score_col = 'cosine_score' if 'cosine_score' in df.columns else ('余弦相似度' if '余弦相似度' in df.columns else None)
        
        if method_col and rank_col:
            df = df.sort_values([sort_col, method_col, rank_col])
        elif rank_col:
            df = df.sort_values([sort_col, rank_col])
        elif score_col:
            df = df.sort_values([sort_col, score_col], ascending=[True, False])
        else:
            df = df.sort_values(sort_col)
        df = df.reset_index(drop=True)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Results')
        worksheet = writer.sheets['Results']
        
        # 1. 设置列宽
        for idx, col in enumerate(df.columns, 1):
            max_length = len(str(col))
            sample_data = df[col].astype(str).head(100)
            for val in sample_data:
                val_len = len(str(val))
                if val_len > max_length:
                    max_length = val_len
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
        
        # 2. 冻结首行
        worksheet.freeze_panes = 'A2'
        
        # 3. 启用自动筛选
        if len(df) > 0:
            worksheet.auto_filter.ref = worksheet.dimensions
        
        # 4. 设置首行样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 5. 设置自动换行和边框
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=2):
            for idx, cell in enumerate(row, 1):
                if cell.value is not None:
                    cell.border = thin_border
                    if idx in wrap_col_indices:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                    elif isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # 6. 合并样品化合物列的连续相同值单元格（支持英文query_name和中文样品化合物）
        query_col_name = None
        if 'query_name' in df.columns:
            query_col_name = 'query_name'
        elif '样品化合物' in df.columns:
            query_col_name = '样品化合物'
        
        if query_col_name:
            query_col_idx = df.columns.get_loc(query_col_name) + 1
            merge_start = 2
            prev_value = None
            
            for row_idx, value in enumerate(df[query_col_name], start=2):
                if prev_value is None:
                    prev_value = value
                    merge_start = row_idx
                elif value != prev_value:
                    if row_idx - 1 > merge_start:
                        worksheet.merge_cells(
                            start_row=merge_start, start_column=query_col_idx,
                            end_row=row_idx - 1, end_column=query_col_idx
                        )
                        worksheet.cell(row=merge_start, column=query_col_idx).alignment = Alignment(
                            horizontal='left', vertical='center'
                        )
                    prev_value = value
                    merge_start = row_idx
            
            if len(df) > 0 and merge_start < len(df) + 2:
                worksheet.merge_cells(
                    start_row=merge_start, start_column=query_col_idx,
                    end_row=len(df) + 1, end_column=query_col_idx
                )
                worksheet.cell(row=merge_start, column=query_col_idx).alignment = Alignment(
                    horizontal='left', vertical='center'
                )


def process_final_results(output_dir: str, l1_csv: str, l2_csv: str, l3_csv: str, 
                         output_csv: str, l4a_csv: str = None, l4b_csv: str = None) -> bool:
    """
    处理最终结果汇总（L1 + L2 + L3 + L4a + L4b）
    """
    print(f"\n{'='*60}")
    print("最终结果汇总 (L1 + L2 + L3 + L4a + L4b)")
    print(f"{'='*60}")
    
    all_results = []
    counts = {}
    
    for name, csv_path in [('L1', l1_csv), ('L2', l2_csv), ('L3', l3_csv), ('L4a', l4a_csv), ('L4b', l4b_csv)]:
        if csv_path and os.path.exists(csv_path):
            df = pd.read_csv(csv_path)
            df['identification_level'] = name
            all_results.append(df)
            counts[name] = len(df)
    
    if not all_results:
        print("[ERROR] 没有可用的结果文件")
        return False
    
    final_df = pd.concat(all_results, ignore_index=True)
    
    # 简洁输出：L1: 678 | L2: 123 | L3: 456 | 合并: 1257 条
    count_str = " | ".join(f"{k}: {v}" for k, v in counts.items())
    print(f"  {count_str} | 合并: {len(final_df)} 条")
    
    final_df.to_csv(output_csv, index=False, encoding='utf-8')
    
    format_excel_output(final_df, output_csv.replace('.csv', '.xlsx'),
                       wrap_columns="matched_name,matched_ontology,matched_fragments")
    
    return True


def parse_msp_to_dict(msp_path):
    """解析MSP文件为字典列表"""
    spectra = []
    current = {}
    peaks = []
    
    with open(msp_path, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            line = line.strip()
            if not line:
                # 空行表示一个谱图结束
                if current and peaks:
                    current['peaks'] = peaks
                    spectra.append(current)
                current = {}
                peaks = []
                continue
            
            if ':' in line and not line[0].isdigit():
                key, value = line.split(':', 1)
                key = key.strip().lower()
                value = value.strip()
                
                if key in ['name', 'compound_name']:
                    current['name'] = value
                elif key in ['precursormz', 'precursor_mz', 'pepmass', 'precursor_mass']:
                    try:
                        # PEPMASS 可能带空格后的charge，如 "123.45 1+"
                        value = value.split()[0] if ' ' in value else value
                        current['precursor_mz'] = float(value)
                    except:
                        current['precursor_mz'] = value
                elif key in ['precursortype', 'precursor_type', 'adduct']:
                    current['precursor_type'] = value
                    current['adduct'] = value
                elif key == 'ionmode':
                    current['ion_mode'] = value
                elif key == 'formula':
                    current['formula'] = value
                elif key == 'smiles':
                    current['smiles'] = value
                elif key == 'inchikey':
                    current['inchikey'] = value
                elif key == 'retentiontime':
                    try:
                        current['retention_time'] = float(value)
                    except:
                        current['retention_time'] = value
                elif key == 'ccs':
                    try:
                        current['ccs'] = float(value)
                    except:
                        current['ccs'] = value
                elif key == 'comment':
                    current['comment'] = value
            elif line[0].isdigit() or (line[0] == '-' and len(line) > 1 and line[1].isdigit()):
                # 碎片峰行
                parts = line.split()
                if len(parts) >= 2:
                    try:
                        mz = float(parts[0])
                        intensity = float(parts[1])
                        peaks.append((mz, intensity))
                    except:
                        pass
        
        # 处理最后一个谱图
        if current and peaks:
            current['peaks'] = peaks
            spectra.append(current)
    
    return spectra


def generate_unidentified_msp(identified_df, sample_msp, sample_csv, output_msp, level="L1",
                              additional_identified_csv=None):
    """
    生成未鉴定化合物的MSP文件
    
    参数:
        identified_df: 已鉴定结果DataFrame
        sample_msp: 原始样品MSP文件路径
        sample_csv: 原始样品CSV文件路径
        output_msp: 输出未鉴定MSP文件路径
        level: 鉴定层级
        additional_identified_csv: 额外已鉴定结果CSV路径（L2排除L1已鉴定样品）
    """
    print(f"[INFO] 生成{level}未鉴定MSP文件...")
    
    try:
        # 1. 获取已鉴定的query_name列表
        identified_names = set()
        if 'query_name' in identified_df.columns:
            identified_names = set(identified_df['query_name'].dropna().unique())
        
        print(f"[INFO] {level}已鉴定化合物: {len(identified_names)} 个")
        
        # 1b. 加载额外已鉴定结果（例如L2生成unidentified时，需要同时排除L1已鉴定的样品）
        if additional_identified_csv and os.path.exists(additional_identified_csv):
            try:
                additional_df = pd.read_csv(additional_identified_csv)
                if 'query_name' in additional_df.columns:
                    additional_names = set(additional_df['query_name'].dropna().unique())
                    print(f"[INFO] 额外排除已鉴定化合物（来自{os.path.basename(additional_identified_csv)}）: {len(additional_names)} 个")
                    identified_names = identified_names | additional_names
            except Exception as e:
                print(f"[WARNING] 加载额外已鉴定CSV失败: {e}")
        
        print(f"[INFO] 总排除已鉴定化合物: {len(identified_names)} 个")
        
        # 2. 解析原始MSP文件
        all_spectra = parse_msp_to_dict(sample_msp)
        print(f"[INFO] 原始MSP共 {len(all_spectra)} 条谱图")
        
        # 3. 筛选未鉴定的谱图
        unidentified_spectra = []
        for spec in all_spectra:
            spec_name = spec.get('name', '')
            # 检查是否已鉴定（支持部分匹配）
            is_identified = False
            for identified_name in identified_names:
                if spec_name in identified_name or identified_name in spec_name:
                    is_identified = True
                    break
            
            if not is_identified:
                unidentified_spectra.append(spec)
        
        print(f"[INFO] 未鉴定谱图: {len(unidentified_spectra)} 条")
        
        # 4. 写入未鉴定MSP文件
        if unidentified_spectra:
            with open(output_msp, 'w', encoding='utf-8') as f:
                for spec in unidentified_spectra:
                    f.write(f"NAME: {spec.get('name', 'Unknown')}\n")
                    if 'precursor_mz' in spec:
                        f.write(f"PRECURSORMZ: {spec['precursor_mz']}\n")
                    if 'precursor_type' in spec:
                        f.write(f"PRECURSORTYPE: {spec['precursor_type']}\n")
                    if 'adduct' in spec:
                        f.write(f"ADDUCT: {spec['adduct']}\n")
                    if 'ion_mode' in spec:
                        f.write(f"IONMODE: {spec['ion_mode']}\n")
                    if 'formula' in spec:
                        f.write(f"FORMULA: {spec['formula']}\n")
                    if 'smiles' in spec:
                        f.write(f"SMILES: {spec['smiles']}\n")
                    if 'inchikey' in spec:
                        f.write(f"INCHIKEY: {spec['inchikey']}\n")
                    if 'retention_time' in spec:
                        f.write(f"RETENTIONTIME: {spec['retention_time']}\n")
                    if 'ccs' in spec:
                        f.write(f"CCS: {spec['ccs']}\n")
                    if 'comment' in spec:
                        f.write(f"COMMENT: {spec['comment']}\n")
                    
                    # 写入碎片峰
                    peaks = spec.get('peaks', [])
                    if peaks:
                        f.write(f"Num Peaks: {len(peaks)}\n")
                        for mz, intensity in peaks:
                            f.write(f"{mz}\t{intensity}\n")
                    f.write("\n")
            
            print(f"[INFO] 未鉴定MSP已保存: {output_msp}")
        else:
            # 创建空文件
            with open(output_msp, 'w', encoding='utf-8') as f:
                pass
            print(f"[INFO] 所有化合物均已鉴定，创建空MSP文件: {output_msp}")
        
    except Exception as e:
        print(f"[ERROR] 生成未鉴定MSP失败: {e}")
        import traceback
        traceback.print_exc()


def main():
    """主函数"""
    args = parse_arguments()
    
    if args.mode in ['L1', 'L2', 'L3', 'L4']:
        success = process_level_results(args.input, args.output, args.mode,
                                       args.sample_msp, args.sample_csv,
                                       mc_input=args.mc_input,
                                       dreams_input=args.dreams_input,
                                       additional_identified_csv=getattr(args, 'additional_identified_csv', None),
                                       skip_excel=getattr(args, 'skip_excel', False))
    elif args.mode == 'final':
        success = process_final_results(
            args.input,  # 作为output_dir
            args.l1_results,
            args.l2_results,
            args.l3_results,
            args.output,
            l4a_csv=args.l4a_results,
            l4b_csv=args.l4b_results,
        )
    else:
        print(f"[ERROR] 未知模式: {args.mode}")
        return 1
    
    return 0 if success else 1


if __name__ == "__main__":
    sys.exit(main())
