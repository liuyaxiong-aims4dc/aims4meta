#!/usr/bin/env python3
"""
多层鉴定结果汇总工具

统一处理：
- 各层（L1/L2/L3/L4）结果汇总
- 最终结果汇总（L1+L2+L3+L4）
- 格式化Excel输出
"""

import argparse
import os
import re
import sys
import pandas as pd
import numpy as np
from pathlib import Path
from typing import List, Optional, Dict

# 单同位素精确质量（Da），用于从分子式计算理论 m/z
_MONOISOTOPIC_MASS = {
    'H': 1.007825, 'C': 12.000000, 'N': 14.003074, 'O': 15.994915,
    'S': 31.972071, 'P': 30.973762, 'F': 18.998403, 'Cl': 34.968853,
    'Br': 78.918338, 'I': 126.904468, 'Na': 22.989770, 'K': 38.963707,
}


def _exact_mass_from_formula(formula: str) -> float:
    """从分子式（如 C20H28N2O14）计算单同位素精确质量"""
    mass = 0.0
    for m in re.finditer(r'([A-Z][a-z]?)(\d*)', str(formula).strip()):
        elem, count = m.group(1), m.group(2)
        count = int(count) if count else 1
        if elem in _MONOISOTOPIC_MASS:
            mass += _MONOISOTOPIC_MASS[elem] * count
    return mass


def parse_arguments():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(description="多层鉴定结果汇总")
    parser.add_argument("--mode", required=True,
                       choices=['L1', 'L2', 'L3', 'final', 'final_excel', 'generate_unidentified'],
                       help="汇总模式: L1/L2/L3(SIRIUS结构鉴定)/final/final_excel(仅生成Excel)/generate_unidentified")
    parser.add_argument("--input", default=None,
                       help="输入CSV文件路径（L1/L2/L3模式）或结果目录（final模式）")
    parser.add_argument("--mc_input", default=None, help="MC鉴定结果CSV路径（L1/L2模式，与--dreams_input搭配使用）")
    parser.add_argument("--dreams_input", default=None, help="DreaMS鉴定结果CSV路径（L1/L2模式，与--mc_input搭配使用）")
    parser.add_argument("--output", required=True, help="输出文件路径")
    parser.add_argument("--l1_results", help="L1结果CSV路径（final模式）")
    parser.add_argument("--l2_results", help="L2结果CSV路径（final模式）")
    parser.add_argument("--l3_results", help="L3结果CSV路径（final模式）")
    parser.add_argument("--l4_results", help="L4结果CSV路径（final模式）")
    parser.add_argument("--sample_msp", help="原始样品MSP文件路径（用于generate_unidentified模式和最终汇总Excel文件名）")
    parser.add_argument("--ion_mode", default=None, choices=['POS', 'NEG'], help="离子模式（用于Excel文件名标注）")
    parser.add_argument("--output_msp", help="输出未鉴定MSP文件路径（generate_unidentified模式）")
    parser.add_argument("--sample_csv", help="[已废弃] 原始样品CSV文件路径")
    parser.add_argument("--additional_identified_csv", help="[已废弃] 额外已鉴定结果CSV")
    return parser.parse_args()


def get_column_order(level: str) -> List[str]:
    """
    获取各层的标准列顺序
    
    参数:
        level: 鉴定层级 (L1/L2/L3)
    
    返回:
        标准列顺序列表
    """
    # 基础列顺序（所有层级通用）
    base_order = [
        # 1. 查询化合物信息
        'query_name',
        # 2. 鉴定化合物基本信息
        'matched_name', 'matched_smiles', 'matched_inchikey', 'matched_formula',
        # 3. 鉴定质量指标
        'cosine_score', 'matched_peaks_ratio', 'matched_fragments',
        # 4. 质谱信息
        'precursor_mass_error', 'adduct',
        # 5. Ontology信息
        'matched_ontology',
        # 6. 来源信息
        'source_method', 'source_database',
        # 7. 原始样本信息（从原始CSV关联）
        'CCS_error',
        'ccs_combined',
        # 8. 其他样本信息
        'Maximum Abundance',
        # 9. 同位素信息
        'isotope_similarity',
        # 10. 综合得分和排名
        'comprehensive_score', 'rank',
    ]
    
    # L3特有列（SIRIUS结构鉴定）
    l3_specific = ['formula', 'structure_confidence', 'sirius_score', 'zodiac_score']
    
    if level == 'L1':
        return base_order
    elif level == 'L2':
        return base_order
    elif level == 'L3':
        # L3 SIRIUS结构鉴定：列顺序不同（无library_precursor_mz，由formula+adduct动态计算）
        return [
            'query_name',
            'matched_name', 'matched_formula', 'matched_inchikey',
            'sirius_score', 'structure_confidence',
            'precursor_mass_error', 'adduct',
            'ccs_combined',
            'isotope_similarity',
            'comprehensive_score', 'rank',
        ]
    elif level == 'L4':
        # L4已改为纯分子网络，不再有单层鉴定结果
        return []
    
    return base_order


def get_column_rename_map(level: str) -> dict:
    """
    获取各层的列名中文映射
    
    参数:
        level: 鉴定层级 (L1/L2/L3)
    
    返回:
        {英文列名: 中文列名} 字典
    """
    # 通用列名映射
    base_map = {
        # 1. 查询化合物信息
        'query_name': '样品化合物',
        # 2. 鉴定化合物基本信息
        'matched_name': '鉴定化合物',
        'matched_smiles': '鉴定SMILES',
        'matched_inchikey': '鉴定InChIKey',
        'matched_formula': '鉴定分子式',
        # 3. 鉴定质量指标
        'cosine_score': '余弦相似度',
        'matched_peaks_ratio': '匹配碎片数/数据库碎片数',
        'matched_fragments': '鉴定碎片详情',
        # 4. 质谱信息
        'precursor_mz': '样品母离子(m/z)',
        'library_precursor_mz': '鉴定母离子(m/z)',
        'precursor_ppm_diff': '母离子质量偏差(ppm)',
        'precursor_mass_error': '样品/库母离子(m/z)(偏差ppm)',
        'adduct': '加合物类型',
        # L4独有列已移除（纯分子网络不再纳入鉴定结果）
        # 5. Ontology信息
        'matched_ontology': '分类',
        # 6. 来源信息
        'source_method': '鉴定方法',
        'source_database': '来源数据库',
        'comprehensive_score': '综合得分',
        'rank': '排名',
        # 7. 原始样本信息
        'Retention time (min)': '保留时间(min)',
        'ccs_combined': '实测/预测CCS(Å²)(偏差%)',
        'Maximum Abundance': '最大丰度',
        # 8. 同位素信息
        'isotope_similarity': '同位素相似度',
        # 9. 其他
        'matched_name_cn': '鉴定化合物(中文)',
    }
    
    # L2特有列
    l2_map = {
        'simulated_library': '模拟库',
        'source_tool': '预测工具',
    }

    # L3特有列（SIRIUS结构鉴定）
    l3_map = {
        'sirius_score': 'SIRIUS得分',
        'structure_confidence': '结构置信度',
    }

    # 最终汇总特有列
    final_map = {
        'identification_level': '鉴定层级',
    }

    if level == 'L2':
        base_map.update(l2_map)
    elif level == 'L3':
        base_map.update(l3_map)
    elif level == 'final':
        base_map.update(l3_map)
        base_map.update(final_map)

    return base_map


def get_columns_to_drop(level: str) -> List[str]:
    """
    获取各层需要删除的列
    
    参数:
        level: 鉴定层级 (L1/L2/L3)
    
    返回:
        需要删除的列名列表
    """
    # 所有层级都不需要的列
    common_drop = [
        'sample_charge',
        'isotope_score',
        'library_ccs',
        'prediction_source',
        # Isotope Distribution 保留，同位素相似度计算需要
        # 保留时间列不需要（第一列已有相关信息）
        'Retention time (min)',
        # 原始CSV中不需要保留的列
        'Neutral mass (Da)', 'm/z', 'Charge',
        'Chromatographic peak width (min)', 'Identifications',
        'Minimum CV%',
        # 原始CSV中的丰度列（样本特定名称）
        '20260203-LIHUA-HDMSE-POS', '20260203-LIHUA-HDMSE-POS.1',
        # 原始CSV中的已接受信息（冗余）
        'Accepted Compound ID', 'Accepted Description',
        # 原始CSV中的加合物/分子式/得分（我们有自己的matched_版本）
        'Adducts', 'Formula', 'Score', 'Fragmentation Score',
        'Mass Error (ppm)', 'Isotope Similarity',
        'Isotope Distribution',
        'Retention Time Error (mins)', 'dCCS (angstrom^2)', 'Compound Link',
    ]
    
    if level == 'L1':
        return common_drop
    elif level == 'L2':
        return common_drop
    elif level == 'L3':
        # L3 SIRIUS结构鉴定可能没有这些列
        return common_drop + ['library_ccs', 'library_precursor_mz', 'precursor_ppm_diff',
                              'formula', 'zodiac_score', 'precursor_formula']
    elif level == 'L4':
        # L4已改为纯分子网络
        return []
    
    return common_drop


def get_wrap_columns(level: str) -> str:
    """
    获取需要自动换行的列
    
    参数:
        level: 鉴定层级 (L1/L2/L3)
    
    返回:
        逗号分隔的列名字符串
    """
    if level == 'L3':
        return "matched_name,matched_formula"
    else:
        return "matched_name,matched_ontology,matched_fragments"


def merge_mc_dreams(mc_csv: str, dreams_csv: str) -> Optional[pd.DataFrame]:
    """
    合并 MatchMS 和 DreaMS 鉴定结果（并集策略）

    参数:
        mc_csv: MatchMS鉴定结果CSV路径
        dreams_csv: DreaMS鉴定结果CSV路径

    返回:
        合并后的DataFrame，失败返回None
    """
    if mc_csv and os.path.exists(mc_csv):
        try:
            mc_df = pd.read_csv(mc_csv)
            if mc_df.empty:
                mc_df = pd.DataFrame()
            # 列名统一：isotope_score → isotope_similarity
            if 'isotope_score' in mc_df.columns:
                mc_df = mc_df.rename(columns={'isotope_score': 'isotope_similarity'})
            mc_count = len(mc_df)
        except (pd.errors.EmptyDataError, pd.errors.ParserError):
            mc_df = pd.DataFrame()
            mc_count = 0
    else:
        mc_df = pd.DataFrame()
        mc_count = 0
    
    if dreams_csv and os.path.exists(dreams_csv):
        try:
            dreams_df = pd.read_csv(dreams_csv)
            if dreams_df.empty:
                dreams_df = pd.DataFrame()
            dreams_count = len(dreams_df)
        except (pd.errors.EmptyDataError, pd.errors.ParserError):
            dreams_df = pd.DataFrame()
            dreams_count = 0
    else:
        dreams_df = pd.DataFrame()
        dreams_count = 0
    
    if mc_df.empty and dreams_df.empty:
        print("[WARN] MatchMS和DreaMS结果均为空，返回空DataFrame")
        return pd.DataFrame()
    
    # 合并
    if mc_df.empty:
        combined = dreams_df.copy()
    elif dreams_df.empty:
        combined = mc_df.copy()
    else:
        combined = pd.concat([mc_df, dreams_df], ignore_index=True)
    
    # 按 comprehensive_score 排序（如存在）
    if 'comprehensive_score' in combined.columns:
        combined = combined.sort_values('comprehensive_score', ascending=False)
    elif 'cosine_score' in combined.columns:
        combined = combined.sort_values('cosine_score', ascending=False)
    
    # 只删除完全重复的行（所有列都相同），保留所有不同的鉴定结果
    combined = combined.drop_duplicates(keep='first')
    
    combined = combined.reset_index(drop=True)
    
    # 按 source_method 分组重新计算 rank
    if 'query_name' in combined.columns and 'source_method' in combined.columns and 'rank' in combined.columns:
        # 优先使用综合评分排序
        if 'comprehensive_score' in combined.columns:
            combined = combined.sort_values(['query_name', 'source_method', 'comprehensive_score'], ascending=[True, True, False])
        elif 'cosine_score' in combined.columns:
            combined = combined.sort_values(['query_name', 'source_method', 'cosine_score'], ascending=[True, True, False])
        combined['rank'] = combined.groupby(['query_name', 'source_method']).cumcount() + 1
        combined = combined.reset_index(drop=True)
    
    print(f"  MatchMS: {mc_count} + DreaMS: {dreams_count} → 整合: {len(combined)} 条")
    return combined


def process_level_results(input_csv: str, output_csv: str, level: str,
                         sample_msp: str = None, sample_csv: str = None,
                         mc_input: str = None, dreams_input: str = None,
                         additional_identified_csv: str = None) -> bool:
    """
    处理指定层级的鉴定结果（仅输出CSV，Excel由最终汇总统一生成）

    参数:
        input_csv: 输入CSV文件路径（已合并的单文件模式）
        output_csv: 输出CSV文件路径
        level: 鉴定层级 (L1/L2/L3)
        mc_input: MatchMS鉴定结果CSV路径（MatchMS+DreaMS整合模式）
        dreams_input: DreaMS鉴定结果CSV路径（MatchMS+DreaMS整合模式）
        sample_msp/sample_csv/additional_identified_csv: 已废弃，不再使用

    返回:
        是否成功
    """
    print(f"\n{'='*60}")
    print(f"{level} 结果汇总")
    print(f"{'='*60}")
    
    # 1. 加载数据
    if mc_input or dreams_input:
        df = merge_mc_dreams(mc_input, dreams_input)
        if df is None:
            return False
        # 如果结果为空，直接保存空CSV并返回
        if df.empty:
            print("  [WARN] 鉴定结果为空（L1未加载数据库）")
            # 创建带标准列头的空DataFrame
            standard_columns = get_column_order(level)
            df = pd.DataFrame(columns=standard_columns)
            df.to_csv(output_csv, index=False, encoding='utf-8')
            print(f"  输出: {output_csv}")
            return True
    elif input_csv:
        try:
            df = pd.read_csv(input_csv)
            print(f"  加载: {len(df)} 条记录")
            
            # L4已改为纯分子网络，不再有单层鉴定结果CSV
            # 旧版L4列名统一逻辑已移除
            
            # 重新计算排名（按综合评分）
            if 'query_name' in df.columns and 'source_method' in df.columns and 'rank' in df.columns:
                # 优先使用综合评分排序
                if 'comprehensive_score' in df.columns:
                    df = df.sort_values(['query_name', 'source_method', 'comprehensive_score'], ascending=[True, True, False])
                elif 'cosine_score' in df.columns:
                    df = df.sort_values(['query_name', 'source_method', 'cosine_score'], ascending=[True, True, False])
                df['rank'] = df.groupby(['query_name', 'source_method']).cumcount() + 1
                df = df.reset_index(drop=True)
                print("  重新计算排名完成")
        except Exception as e:
            print(f"[ERROR] 加载失败: {e}")
            return False
    else:
        print("[ERROR] 未提供输入")
        return False

    # 2. 合并CCS列为 ccs_combined（实测/预测(偏差%)），始终删除原始 CCS 列
    has_measured = 'CCS (angstrom^2)' in df.columns
    has_predicted = 'predicted_ccs' in df.columns
    if has_measured:
        df['ccs_combined'] = ""
        for idx in df.index:
            measured = df.loc[idx, 'CCS (angstrom^2)']
            predicted = df.loc[idx, 'predicted_ccs'] if has_predicted else ""
            deviation = df.loc[idx, 'predicted_ccs_deviation_pct'] if 'predicted_ccs_deviation_pct' in df.columns else ""

            if pd.notna(measured) and str(measured).strip() not in ('', 'nan'):
                m_val = float(measured)
                if has_predicted and pd.notna(predicted) and str(predicted).strip() not in ('', 'nan'):
                    p_val = float(predicted)
                    if pd.notna(deviation) and str(deviation).strip() not in ('', 'nan'):
                        df.loc[idx, 'ccs_combined'] = f"{m_val:.2f}/{p_val:.2f} ({deviation:+.1f}%)"
                    else:
                        df.loc[idx, 'ccs_combined'] = f"{m_val:.2f}/{p_val:.2f}"
                else:
                    # 仅实测值（无预测CCS）
                    df.loc[idx, 'ccs_combined'] = f"{m_val:.2f}/-"

    # 始终删除原始 CCS 列（信息已合并到 ccs_combined）
    df = df.drop(columns=['CCS (angstrom^2)', 'predicted_ccs', 'predicted_ccs_deviation_pct', 'CCS_error'], errors='ignore')

    # L3: 若 SIRIUS 未提供 library_precursor_mz（旧版输出），从 precursor_formula 或 matched_formula+adduct 兜底计算
    if level == 'L3' and 'matched_formula' in df.columns and 'adduct' in df.columns:
        need_calc = False
        if 'library_precursor_mz' not in df.columns:
            need_calc = True
        else:
            # 有列但全部为空 → 需要兜底计算
            has_nonempty = df['library_precursor_mz'].apply(
                lambda v: pd.notna(v) and str(v).strip() not in ('', 'nan')).any()
            if not has_nonempty:
                need_calc = True

        if need_calc:
            # 首选：从 precursor_formula 算精确质量（SIRIUS 已将加合物原子融入前体化学式）
            has_precursor_formula = ('precursor_formula' in df.columns and
                df['precursor_formula'].apply(lambda v: pd.notna(v) and str(v).strip() not in ('', 'nan')).any())

            if has_precursor_formula:
                for idx in df.index:
                    pf = df.loc[idx, 'precursor_formula']
                    if pd.notna(pf) and str(pf).strip():
                        try:
                            # 去掉末尾的电荷符号（如 "C20H17O9-" → "C20H17O9"）
                            library_mz = round(_exact_mass_from_formula(str(pf).strip().rstrip('+-')), 4)
                            df.loc[idx, 'library_precursor_mz'] = library_mz

                            sample_mz = df.loc[idx, 'precursor_mz']
                            if pd.notna(sample_mz) and str(sample_mz).strip():
                                ppm = (float(sample_mz) - library_mz) / library_mz * 1e6
                                df.loc[idx, 'precursor_ppm_diff'] = ppm
                        except Exception:
                            pass
            else:
                # 兜底：从 matched_formula + adduct map 计算（兼容旧版无 precursor_formula 列）
                adduct_mass_map = {
                    '[M-H]-': -1.007825, '[M+H]+': 1.007825,
                    '[M+Na]+': 22.989770, '[M+K]+': 38.963707,
                    '[M+NH4]+': 18.034374, '[M+Cl]-': 34.968853,
                    '[M+FA-H]-': 44.997654, '[M-H2O-H]-': -19.018390,
                    '[M+HAc-H]-': 59.013305,
                }
                for idx in df.index:
                    formula = df.loc[idx, 'matched_formula']
                    adduct = df.loc[idx, 'adduct']
                    if pd.notna(formula) and str(formula).strip() and pd.notna(adduct) and str(adduct).strip():
                        try:
                            exact_mass = _exact_mass_from_formula(str(formula).strip())
                            adduct_delta = adduct_mass_map.get(str(adduct).strip().replace(' ', ''), 0)
                            library_mz = round(exact_mass + adduct_delta, 4)
                            df.loc[idx, 'library_precursor_mz'] = library_mz

                            sample_mz = df.loc[idx, 'precursor_mz']
                            if pd.notna(sample_mz) and str(sample_mz).strip():
                                ppm = (float(sample_mz) - library_mz) / library_mz * 1e6
                                df.loc[idx, 'precursor_ppm_diff'] = ppm
                        except Exception:
                            pass

    # 合并母离子列
    if 'precursor_mz' in df.columns and 'library_precursor_mz' in df.columns:
        df['precursor_mass_error'] = ""
        for idx in df.index:
            sample_mz = df.loc[idx, 'precursor_mz']
            lib_mz = df.loc[idx, 'library_precursor_mz']
            ppm = df.loc[idx, 'precursor_ppm_diff'] if 'precursor_ppm_diff' in df.columns else ""

            if pd.notna(sample_mz) and sample_mz != "":
                if pd.notna(lib_mz) and lib_mz != "" and pd.notna(ppm) and ppm != "":
                    df.loc[idx, 'precursor_mass_error'] = f"{sample_mz}/{lib_mz}({ppm:+.1f}ppm)"
                else:
                    df.loc[idx, 'precursor_mass_error'] = f"{sample_mz}/-"

        df = df.drop(columns=['precursor_mz', 'library_precursor_mz', 'precursor_ppm_diff'], errors='ignore')

    # 3. 优化列顺序
    column_order = get_column_order(level)
    columns_to_drop = get_columns_to_drop(level)
    wrap_columns = get_wrap_columns(level)
    
    existing_columns = [col for col in column_order if col in df.columns]
    other_columns = [col for col in df.columns if col not in column_order]
    other_columns = [col for col in other_columns if col not in columns_to_drop]
    
    final_columns = existing_columns + other_columns
    df = df[final_columns]
    
    # 3. 保存CSV（Excel由最终汇总统一生成，各层不再单独生成）
    try:
        df.to_csv(output_csv, index=False, encoding='utf-8')
    except Exception as e:
        print(f"[ERROR] CSV保存失败: {e}")
        return False
    
    print(f"  输出: {output_csv} ({len(df)} 条)")
    
    return True


def get_sheet_name_from_msp(msp_path: str = None) -> str:
    """
    从MSP文件路径提取sheet名称
    
    参数:
        msp_path: MSP文件路径
        
    返回:
        适合作为Excel sheet名称的字符串（最长31字符）
    """
    if not msp_path:
        return "Results"
    
    msp_filename = os.path.basename(msp_path)
    name_without_ext = os.path.splitext(msp_filename)[0]
    
    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
    for char in invalid_chars:
        name_without_ext = name_without_ext.replace(char, '_')
    
    if len(name_without_ext) > 31:
        name_without_ext = name_without_ext[:31]
    
    return name_without_ext or "Results"


def get_excel_filename_from_msp(msp_path: str = None, output_csv: str = None, level: str = "", ion_mode: str = None) -> str:
    """
    从MSP文件路径生成Excel文件名
    
    参数:
        msp_path: MSP文件路径
        output_csv: 原始输出CSV路径
        level: 层级（L1, L2, L3, L4, final）
        ion_mode: 离子模式（POS/NEG），用于文件名标注
        
    返回:
        包含MSP文件名和离子模式的Excel文件路径
    """
    if not msp_path:
        base = output_csv.replace('.csv', '.xlsx') if output_csv else output_csv
        # 无MSP但有ion_mode时也插入
        if ion_mode and output_csv:
            base_name = os.path.basename(base)
            dir_name = os.path.dirname(base)
            name_no_ext = os.path.splitext(base_name)[0]
            new_name = f"{name_no_ext}({ion_mode}).xlsx"
            base = os.path.join(dir_name, new_name)
        return base
    
    msp_filename = os.path.basename(msp_path)
    name_without_ext = os.path.splitext(msp_filename)[0]
    
    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
    for char in invalid_chars:
        name_without_ext = name_without_ext.replace(char, '_')
    
    if output_csv:
        output_dir = os.path.dirname(output_csv)
        base_name = os.path.basename(output_csv).replace('.csv', '')
        # 插入离子模式标注
        if ion_mode:
            new_filename = f"{name_without_ext}({ion_mode})_{base_name}.xlsx"
        else:
            new_filename = f"{name_without_ext}_{base_name}.xlsx"
        return os.path.join(output_dir, new_filename)
    
    return output_csv


def format_excel_output(df: pd.DataFrame, output_path: str, 
                       wrap_columns: str = "",
                       sheet_name: str = "Results"):
    """
    将DataFrame格式化为Excel文件（内置功能）
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return
    
    # 确保输出路径是.xlsx
    if not output_path.endswith('.xlsx'):
        output_path = output_path.replace('.csv', '.xlsx')
        if not output_path.endswith('.xlsx'):
            output_path += '.xlsx'
    
    # 解析需要换行的列
    wrap_cols = [c.strip() for c in wrap_columns.split(',') if c.strip()]
    wrap_col_indices = []
    for col_name in wrap_cols:
        if col_name in df.columns:
            col_idx = df.columns.get_loc(col_name) + 1
            wrap_col_indices.append(col_idx)
    
    # 按样品化合物排序（支持英文query_name和中文样品化合物）
    sort_col = None
    if 'query_name' in df.columns:
        sort_col = 'query_name'
    elif '样品化合物' in df.columns:
        sort_col = '样品化合物'
    
    if sort_col:
        # 检查可用的辅助排序列（支持中英文列名）
        rank_col = 'rank' if 'rank' in df.columns else ('排名' if '排名' in df.columns else None)
        method_col = 'source_method' if 'source_method' in df.columns else ('鉴定方法' if '鉴定方法' in df.columns else None)
        score_col = 'cosine_score' if 'cosine_score' in df.columns else ('余弦相似度' if '余弦相似度' in df.columns else None)
        
        if method_col and rank_col:
            df = df.sort_values([sort_col, method_col, rank_col])
        elif rank_col:
            df = df.sort_values([sort_col, rank_col])
        elif score_col:
            df = df.sort_values([sort_col, score_col], ascending=[True, False])
        else:
            df = df.sort_values(sort_col)
        df = df.reset_index(drop=True)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        
        # 1. 设置列宽
        wide_columns = {'鉴定碎片详情', 'matched_fragments'}  # 内容较长的列放宽上限
        for idx, col in enumerate(df.columns, 1):
            max_length = len(str(col))
            sample_data = df[col].astype(str).head(100)
            for val in sample_data:
                val_len = len(str(val))
                if val_len > max_length:
                    max_length = val_len
            cap = 80 if col in wide_columns else 50
            min_width = 40 if col in wide_columns else 0
            adjusted_width = min(max_length + 2, cap)
            adjusted_width = max(adjusted_width, min_width)
            worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
        
        # 2. 冻结首行
        worksheet.freeze_panes = 'A2'
        
        # 3. 启用自动筛选
        if len(df) > 0:
            worksheet.auto_filter.ref = worksheet.dimensions
        
        # 4. 设置首行样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 5. 设置自动换行和边框
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=2):
            for idx, cell in enumerate(row, 1):
                if cell.value is not None:
                    cell.border = thin_border
                    if idx in wrap_col_indices:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                    elif isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')

        # 6. 按 query_name 合并共享列的连续相同值单元格
        _merge_shared_cells(worksheet, df, wrap_col_indices)

        
def _merge_shared_cells(worksheet, df: pd.DataFrame, wrap_col_indices: set):
    """将同一 query_name 的连续行在共享列上合并单元格

    共享列（样品级信息，不随候选变化）：
      - 样品化合物 / query_name
      - 加合物类型 / adduct
      - 实测/预测CCS(Å²)(偏差%) / ccs_combined
      - 最大丰度 / Maximum Abundance
    """
    # 找到 query_name 列（中/英文名均可）
    query_col_idx = None
    for idx, col in enumerate(df.columns, 1):
        if col in ('样品化合物', 'query_name'):
            query_col_idx = idx
            break
    if query_col_idx is None:
        return

    # 识别共享列索引（排除候选特有列 + 换行列）
    candidate_cols_pattern = re.compile(
        r'^(matched_|cosine_|余弦|匹配碎片|鉴定碎片|鉴定SMILES|鉴定InChIKey|鉴定分子式|'
        r'鉴定化合物|分类|鉴定方法|来源数据库|综合得分|排名|simulated_library|'
        r'source_tool|SIRIUS|结构置信度|isomer_id|'
        r'source_(method|database)|isotope_similarity|同位素相似度|'
        r'library_precursor|鉴定母离子|precursor_mass_error|母离子.*偏差|'
        r'identification_level|鉴定层级|rank|comprehensive_score|'
        r'sample_charge|isotope_score|library_ccs|prediction_source)'
    )
    shared_col_indices = set()
    for idx, col in enumerate(df.columns, 1):
        if idx == query_col_idx:
            shared_col_indices.add(idx)
            continue
        if idx in wrap_col_indices:
            continue  # 换行列内容长，不合并
        if candidate_cols_pattern.search(col):
            continue
        shared_col_indices.add(idx)

    if not shared_col_indices:
        return

    # 从 Excel 第 2 行开始扫描，按 query_col 的值分组
    from openpyxl.utils import get_column_letter
    total_rows = len(df) + 1  # +1 for header
    row = 2
    while row <= total_rows:
        cell_val = worksheet.cell(row=row, column=query_col_idx).value
        # 跳过空值行（被 collapse 置空的后续行）
        if cell_val is None or str(cell_val).strip() == '':
            row += 1
            continue
        # 找到连续相同 query_name 且值非空的区间
        end_row = row
        while end_row + 1 <= total_rows:
            next_val = worksheet.cell(row=end_row + 1, column=query_col_idx).value
            if next_val == cell_val:
                end_row += 1
            else:
                break
        if end_row > row:
            for col_idx in shared_col_indices:
                col_letter = get_column_letter(col_idx)
                try:
                    worksheet.merge_cells(f'{col_letter}{row}:{col_letter}{end_row}')
                except Exception:
                    pass
        row = end_row + 1


def collapse_shared_query_columns(df: pd.DataFrame, query_col: str = 'query_name') -> pd.DataFrame:
    """将同一 query_name 多行中的共享列（样品级信息）合并去重

    对于每个 query_name 组，首行保留全部值，后续行将共享列置空。
    候选特有列（matched_name, cosine_score, 碎片详情等）保持每行独立。

    返回:
        处理后的 DataFrame（新增列的顺序不变）
    """
    if query_col not in df.columns:
        return df

    # 候选特有列正则：这些列在每行独立显示，不参与合并
    candidate_cols_pattern = re.compile(
        r'^(matched_|cosine_|余弦|匹配碎片|鉴定碎片|鉴定SMILES|鉴定InChIKey|鉴定分子式|'
        r'鉴定化合物|分类|鉴定方法|来源数据库|综合得分|排名|simulated_library|'
        r'source_tool|SIRIUS|结构置信度|isomer_id|'
        r'source_(method|database)|isotope_similarity|同位素相似度|'
        r'library_precursor|鉴定母离子|precursor_mass_error|母离子.*偏差|'
        r'identification_level|鉴定层级|rank|comprehensive_score)'
    )

    df = df.copy()
    grouped = df.groupby(query_col, sort=False)
    for _, indices in grouped.groups.items():
        if len(indices) <= 1:
            continue
        # 首行保留全部，后续行只保留候选特有列 + query_col 本身（用于 merge_cells 检测）
        first_idx = indices[0]
        for idx in indices[1:]:
            for col in df.columns:
                if col == query_col:
                    continue  # 保留 query_col 值，供 _merge_shared_cells 合并
                # 候选特有列保留原值
                if candidate_cols_pattern.search(col):
                    continue
                # 共享列置空（合并后 Excel 只显示首行值）
                df.at[idx, col] = '' if isinstance(df.at[first_idx, col], str) else pd.NA

    return df


def process_final_results(output_dir: str, l1_csv: str, l2_csv: str, l3_csv: str,
                         output_csv: str, l4_csv: str = None,
                         sample_msp: str = None, ion_mode: str = None) -> bool:
    """
    处理最终结果汇总（L1 + L2 + L3）
    L4为分子网络，不纳入鉴定结果汇总
    """
    print(f"\n{'='*60}")
    print("最终结果汇总 (L1 + L2 + L3)")
    print(f"{'='*60}")

    # 尝试从样品CSV提取实测CCS
    measured_ccs_dict = {}
    if sample_msp:
        sample_csv = sample_msp.replace('.msp', '.csv')
        if os.path.exists(sample_csv):
            try:
                import re
                sample_df = pd.read_csv(sample_csv, skiprows=2)
                if 'Compound' in sample_df.columns and 'CCS (angstrom^2)' in sample_df.columns:
                    sample_ccs = sample_df[['Compound', 'CCS (angstrom^2)']].copy()
                    sample_ccs.columns = ['compound_id', 'measured_ccs']
                    sample_ccs['measured_ccs'] = pd.to_numeric(sample_ccs['measured_ccs'], errors='coerce')

                    # 从compound_id提取RT和m/z: "1.13_448.1722n" -> RT=1.13, mz=448.1722
                    def parse_compound_id(cid):
                        match = re.match(r'([\d.]+)_([\d.]+)[mn]', str(cid))
                        if match:
                            return float(match.group(1)), float(match.group(2))
                        return None, None

                    sample_ccs['rt'], sample_ccs['mz'] = zip(*sample_ccs['compound_id'].apply(parse_compound_id))

                    # 创建查找字典: (rt, mz) -> measured_ccs
                    for _, row in sample_ccs.iterrows():
                        if pd.notna(row['rt']) and pd.notna(row['mz']) and pd.notna(row['measured_ccs']):
                            key = (round(row['rt'], 2), round(row['mz'], 4))
                            measured_ccs_dict[key] = row['measured_ccs']

                    print(f"从样品CSV提取实测CCS: {len(measured_ccs_dict)}条")
            except Exception as e:
                print(f"警告: 无法从样品CSV提取实测CCS: {e}")

    all_results = []
    counts = {}
    
    for name, csv_path in [('L1_实验数据库', l1_csv), ('L2_模拟数据库', l2_csv), ('L3_SIRIUS结构鉴定', l3_csv)]:
        if csv_path and os.path.exists(csv_path):
            try:
                df = pd.read_csv(csv_path)
            except (pd.errors.EmptyDataError, pd.errors.ParserError):
                print(f"跳过空或损坏的结果文件: {name} ({csv_path})")
                continue
            if df.empty:
                print(f"跳过空结果: {name}")
                continue
            df['identification_level'] = name

            # L4类似物筛查已移除（纯分子网络不再纳入鉴定结果）

            # 统一L3列名和结构
            if name.startswith('L3'):
                # 检测L3文件格式（旧版 vs 新版）
                if 'query_precursor_mz' in df.columns and 'matched_precursor_type' in df.columns:
                    # 旧版L3格式
                    l3_rename = {
                        'query_precursor_mz': 'precursor_mz',
                        'matched_precursor_type': 'adduct',
                    }
                    df = df.rename(columns=l3_rename)

                    # 旧版L3的matched_formula是分子式（不再复制到formula列，终稿由matched_formula直接展示）
                    pass
                else:
                    # 新版L3格式：precursor_mz, adduct已经是标准列名
                    # matched_formula是分子式（不再复制到formula列）
                    pass

                # 2. 添加必要列
                if 'source_method' not in df.columns:
                    df['source_method'] = 'SIRIUS'
                if 'source_database' not in df.columns:
                    # 从identification_source提取数据库信息
                    if 'identification_source' in df.columns:
                        df['source_database'] = df['identification_source']
                    else:
                        df['source_database'] = 'BIO+PUBCHEM'
                if 'rank' not in df.columns:
                    df['rank'] = 1

                # 3. 合并置信度列（保留structure_confidence作为主列）
                # csi_score, confidence_exact, confidence_approx 已经在structure_confidence中体现

                # 4. 删除冗余列
                cols_to_drop = ['csi_score', 'confidence_exact', 'confidence_approx', 'identification_source',
                               'formula', 'zodiac_score']
                df = df.drop(columns=[c for c in cols_to_drop if c in df.columns])

                # 5. L3无类似物验证标记（原L4a逻辑已移除）

            all_results.append(df)
            counts[name] = len(df)
    
    if not all_results:
        print("[ERROR] 没有可用的结果文件")
        return False
    
    final_df = pd.concat(all_results, ignore_index=True)
    
    # 清理残余异名列（旧版CSV兼容）
    for old, new in [('cosine_similarity', 'cosine_score'),
                      ('query_precursor_mz', 'precursor_mz'),
                      ('matched_precursor_mz', 'library_precursor_mz')]:
        if old in final_df.columns and new in final_df.columns:
            final_df[new] = final_df[new].fillna(final_df[old])
            final_df = final_df.drop(columns=[old])
        elif old in final_df.columns:
            final_df = final_df.rename(columns={old: new})
    
    # 简洁输出：L1: 678 | L2: 123 | L3: 456 | 合并: 1257 条
    count_str = " | ".join(f"{k}: {v}" for k, v in counts.items())
    print(f"  {count_str} | 合并: {len(final_df)} 条")

    # 匹配实测CCS并计算偏差
    if measured_ccs_dict and 'query_name' in final_df.columns:
        import re

        def extract_rt_mz_from_query_name(query_name):
            """从query_name提取RT和m/z: "Unknown (1.09_318.1499m/z)" -> (1.09, 318.1499)"""
            match = re.search(r'\(([\d.]+)_([\d.]+)[mn]/z\)', str(query_name))
            if match:
                return float(match.group(1)), float(match.group(2))
            return None, None

        def find_measured_ccs(query_name):
            """查找实测CCS"""
            rt, mz = extract_rt_mz_from_query_name(query_name)
            if rt is not None and mz is not None:
                key = (round(rt, 2), round(mz, 4))
                return measured_ccs_dict.get(key)
            return None

        final_df['measured_ccs'] = final_df['query_name'].apply(find_measured_ccs)

        # 计算 CCS 组合列（基于样品 CSV 匹配的实测 CCS）
        # 注意：如各层结果已有 ccs_combined（来自 process_level_results），优先保留
        existing_ccs = final_df.get('ccs_combined', pd.Series(dtype=str)) if 'ccs_combined' in final_df.columns else None

        def calculate_ccs_combined(row):
            measured = row.get('measured_ccs')
            predicted = row.get('predicted_ccs')

            if pd.notna(measured) and pd.notna(predicted):
                deviation = abs(predicted - measured) / measured * 100
                return f"{measured:.2f}/{predicted:.2f} ({deviation:.1f}%)"
            elif pd.notna(predicted):
                return f"-/{predicted:.2f}"
            elif pd.notna(measured):
                return f"{measured:.2f}/-"
            return None

        ccs_from_csv = final_df.apply(calculate_ccs_combined, axis=1)

        if existing_ccs is not None:
            # 优先用各层已有的 ccs_combined，仅空白处用样品 CSV 匹配的填充
            mask_empty = existing_ccs.isna() | existing_ccs.astype(str).str.strip().isin(['', 'nan'])
            final_df['ccs_combined'] = existing_ccs.where(~mask_empty, ccs_from_csv)
        else:
            final_df['ccs_combined'] = ccs_from_csv

        # 删除临时列和旧 CCS 列
        final_df = final_df.drop(columns=['measured_ccs', 'predicted_ccs', 'predicted_ccs_deviation_pct', 'CCS_error'], errors='ignore')

        matched_count = final_df['ccs_combined'].notna().sum()
        print(f"  实测CCS匹配: {matched_count}/{len(final_df)} ({matched_count/len(final_df)*100:.1f}%)")

    # 调整列顺序：identification_level 插到 rank 后面
    cols = list(final_df.columns)
    if 'identification_level' in cols and 'rank' in cols:
        cols.remove('identification_level')
        rank_idx = cols.index('rank')
        cols.insert(rank_idx + 1, 'identification_level')
        final_df = final_df[cols]

    # 加合物类型去空格（SIRIUS 输出 "[M + H]+" → 统一 "[M+H]+"）
    if 'adduct' in final_df.columns:
        final_df['adduct'] = final_df['adduct'].astype(str).str.replace(' ', '', regex=False)

    final_df.to_csv(output_csv, index=False, encoding='utf-8')

    # 同一 query_name 多候选时合并共享列（样品化合物、CCS 等只显示一次）
    final_df = collapse_shared_query_columns(final_df)

    # 中文列名映射（合并所有层级的映射）
    rename_map = get_column_rename_map('final')
    # 添加L2的特殊列映射
    rename_map.update(get_column_rename_map('L2'))

    final_df_cn = final_df.rename(columns=rename_map)
    wrap_columns_cn = [rename_map.get(col, col) for col in "matched_name,matched_ontology,matched_fragments".split(',') if col]
    wrap_columns_str = ','.join(wrap_columns_cn)
    
    final_xlsx = get_excel_filename_from_msp(sample_msp, output_csv, "final", ion_mode=ion_mode)
    sheet_name = get_sheet_name_from_msp(sample_msp)
    format_excel_output(final_df_cn, final_xlsx,
                       wrap_columns=wrap_columns_str,
                       sheet_name=sheet_name)
    
    return True


def parse_msp_to_dict(msp_path):
    """解析MSP文件为字典列表"""
    spectra = []
    current = {}
    peaks = []
    
    with open(msp_path, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            line = line.strip()
            if not line:
                # 空行表示一个谱图结束
                if current and peaks:
                    current['peaks'] = peaks
                    spectra.append(current)
                current = {}
                peaks = []
                continue
            
            if ':' in line and not line[0].isdigit():
                key, value = line.split(':', 1)
                key = key.strip().lower()
                value = value.strip()
                
                if key in ['name', 'compound_name']:
                    current['name'] = value
                elif key in ['precursormz', 'precursor_mz', 'pepmass', 'precursor_mass']:
                    try:
                        # PEPMASS 可能带空格后的charge，如 "123.45 1+"
                        value = value.split()[0] if ' ' in value else value
                        current['precursor_mz'] = float(value)
                    except:
                        current['precursor_mz'] = value
                elif key in ['precursortype', 'precursor_type', 'adduct']:
                    current['precursor_type'] = value
                    current['adduct'] = value
                elif key == 'ionmode':
                    current['ion_mode'] = value
                elif key == 'formula':
                    current['formula'] = value
                elif key == 'smiles':
                    current['smiles'] = value
                elif key == 'inchikey':
                    current['inchikey'] = value
                elif key == 'retentiontime':
                    try:
                        current['retention_time'] = float(value)
                    except:
                        current['retention_time'] = value
                elif key == 'ccs':
                    try:
                        current['ccs'] = float(value)
                    except:
                        current['ccs'] = value
                elif key == 'comment':
                    current['comment'] = value
            elif line[0].isdigit() or (line[0] == '-' and len(line) > 1 and line[1].isdigit()):
                # 碎片峰行
                parts = line.split()
                if len(parts) >= 2:
                    try:
                        mz = float(parts[0])
                        intensity = float(parts[1])
                        peaks.append((mz, intensity))
                    except:
                        pass
        
        # 处理最后一个谱图
        if current and peaks:
            current['peaks'] = peaks
            spectra.append(current)
    
    return spectra


def generate_unidentified_msp(identified_df, sample_msp, sample_csv, output_msp, level="L1",
                              additional_identified_csv=None):
    """
    生成未鉴定化合物的MSP文件
    
    参数:
        identified_df: 已鉴定结果DataFrame
        sample_msp: 原始样品MSP文件路径
        sample_csv: 原始样品CSV文件路径
        output_msp: 输出未鉴定MSP文件路径
        level: 鉴定层级
        additional_identified_csv: 额外已鉴定结果CSV路径（L2排除L1已鉴定样品）
    """
    print(f"[INFO] 生成{level}未鉴定MSP文件...")
    
    try:
        # 1. 获取已鉴定的query_name列表
        identified_names = set()
        if 'query_name' in identified_df.columns:
            identified_names = set(identified_df['query_name'].dropna().unique())
        
        print(f"[INFO] {level}已鉴定化合物: {len(identified_names)} 个")
        
        # 1b. 加载额外已鉴定结果（例如L2生成unidentified时，需要同时排除L1已鉴定的样品）
        if additional_identified_csv and os.path.exists(additional_identified_csv):
            try:
                additional_df = pd.read_csv(additional_identified_csv)
                if 'query_name' in additional_df.columns:
                    additional_names = set(additional_df['query_name'].dropna().unique())
                    print(f"[INFO] 额外排除已鉴定化合物（来自{os.path.basename(additional_identified_csv)}）: {len(additional_names)} 个")
                    identified_names = identified_names | additional_names
            except Exception as e:
                print(f"[WARNING] 加载额外已鉴定CSV失败: {e}")
        
        print(f"[INFO] 总排除已鉴定化合物: {len(identified_names)} 个")
        
        # 2. 解析原始MSP文件
        all_spectra = parse_msp_to_dict(sample_msp)
        print(f"[INFO] 原始MSP共 {len(all_spectra)} 条谱图")
        
        # 3. 筛选未鉴定的谱图
        unidentified_spectra = []
        for spec in all_spectra:
            spec_name = spec.get('name', '')
            # 检查是否已鉴定（支持部分鉴定）
            is_identified = False
            for identified_name in identified_names:
                if spec_name in identified_name or identified_name in spec_name:
                    is_identified = True
                    break
            
            if not is_identified:
                unidentified_spectra.append(spec)
        
        print(f"[INFO] 未鉴定谱图: {len(unidentified_spectra)} 条")
        
        # 4. 写入未鉴定MSP文件
        if unidentified_spectra:
            with open(output_msp, 'w', encoding='utf-8') as f:
                for spec in unidentified_spectra:
                    f.write(f"NAME: {spec.get('name', 'Unknown')}\n")
                    if 'precursor_mz' in spec:
                        f.write(f"PRECURSORMZ: {spec['precursor_mz']}\n")
                    if 'precursor_type' in spec:
                        f.write(f"PRECURSORTYPE: {spec['precursor_type']}\n")
                    if 'adduct' in spec:
                        f.write(f"ADDUCT: {spec['adduct']}\n")
                    if 'ion_mode' in spec:
                        f.write(f"IONMODE: {spec['ion_mode']}\n")
                    if 'formula' in spec:
                        f.write(f"FORMULA: {spec['formula']}\n")
                    if 'smiles' in spec:
                        f.write(f"SMILES: {spec['smiles']}\n")
                    if 'inchikey' in spec:
                        f.write(f"INCHIKEY: {spec['inchikey']}\n")
                    if 'retention_time' in spec:
                        f.write(f"RETENTIONTIME: {spec['retention_time']}\n")
                    if 'ccs' in spec:
                        f.write(f"CCS: {spec['ccs']}\n")
                    if 'comment' in spec:
                        f.write(f"COMMENT: {spec['comment']}\n")
                    
                    # 写入碎片峰
                    peaks = spec.get('peaks', [])
                    if peaks:
                        f.write(f"Num Peaks: {len(peaks)}\n")
                        for mz, intensity in peaks:
                            f.write(f"{mz}\t{intensity}\n")
                    f.write("\n")
            
            print(f"[INFO] 未鉴定MSP已保存: {output_msp}")
        else:
            # 创建空文件
            with open(output_msp, 'w', encoding='utf-8') as f:
                pass
            print(f"[INFO] 所有化合物均已鉴定，创建空MSP文件: {output_msp}")
        
        # 5. 验证输出文件是否确实生成
        if not os.path.exists(output_msp):
            print(f"[ERROR] 未鉴定MSP文件写入后验证失败: {output_msp}")
            return False
        return True
        
    except Exception as e:
        print(f"[ERROR] 生成未鉴定MSP失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """主函数"""
    args = parse_arguments()

    if args.mode == 'generate_unidentified':
        # 生成未鉴定MSP模式
        if not args.input or not args.sample_msp or not args.output_msp:
            print("[ERROR] generate_unidentified模式需要: --input, --sample_msp, --output_msp")
            return 1
        df = pd.read_csv(args.input)
        # 从input路径推断层级（如L2_results.csv → L2）
        level = "L1"
        input_lower = args.input.lower()
        for lv in ["L3", "L2", "L1"]:
            if f"{lv.lower()}_" in input_lower or f"/{lv.lower()}" in input_lower:
                level = lv
                break
        success = generate_unidentified_msp(df, args.sample_msp, None, args.output_msp, level)
        if not success:
            print(f"[ERROR] 生成{level}未鉴定MSP失败")
            return 1
        return 0

    if args.mode in ['L1', 'L2', 'L3']:
        success = process_level_results(args.input, args.output, args.mode,
                                       args.sample_msp, args.sample_csv,
                                       mc_input=args.mc_input,
                                       dreams_input=args.dreams_input,
                                       additional_identified_csv=getattr(args, 'additional_identified_csv', None))
    elif args.mode == 'final_excel':
        # 仅从已有 CSV 生成格式化 Excel（辅助功能已执行完毕后的最终输出）
        if not args.input or not os.path.exists(args.input):
            print(f"[ERROR] final_excel模式需要 --input 为已存在的CSV文件: {args.input}")
            return 1
        df = pd.read_csv(args.input)
        # 合并共享列
        df = collapse_shared_query_columns(df)
        # 中文列名
        rename_map = get_column_rename_map('final')
        rename_map.update(get_column_rename_map('L2'))
        df_cn = df.rename(columns=rename_map)
        wrap_cols = ','.join([rename_map.get(c,c) for c in "matched_name,matched_ontology,matched_fragments".split(',') if c])
        xlsx = get_excel_filename_from_msp(args.sample_msp, args.input, "final", ion_mode=args.ion_mode)
        sheet = get_sheet_name_from_msp(args.sample_msp)
        format_excel_output(df_cn, xlsx, wrap_columns=wrap_cols, sheet_name=sheet)
        print(f"Excel已生成: {xlsx}")
        return 0
    elif args.mode == 'final':
        success = process_final_results(
            args.input,  # 作为output_dir
            args.l1_results,
            args.l2_results,
            args.l3_results,
            args.output,
            l4_csv=args.l4_results,
            sample_msp=args.sample_msp,
            ion_mode=args.ion_mode,
        )
    else:
        print(f"[ERROR] 未知模式: {args.mode}")
        return 1
    
    return 0 if success else 1


if __name__ == "__main__":
    sys.exit(main())
